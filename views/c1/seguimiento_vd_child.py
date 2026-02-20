import streamlit as st
import pandas as pd
import plotly.express as px
import re
from utils.cache_handler import fetch_vd_childs, fetch_carga_childs, fetch_padron
from utils.functions_data import *
from utils.charts import *
from utils.helpers import *
from styles import styles
from datetime import datetime
from constans import *
from st_aggrid import AgGrid, GridOptionsBuilder,JsCode
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, Alignment, Protection
from io import BytesIO
from utils.g_sheet import *

# Nueva vista nominal enfocada en tamizajes y porcentajes
def seguimiento_nominal_tamizajes():
    styles(2)
    colh, colm = st.columns([7,3])
    with colh:
        st.header("📊 Seguimiento Nominal: Tamizajes 6 y 12 meses")
    with colm:
        select_mes = st.selectbox("Mes", ["Jun","Jul","Ago","Set","Oct","Nov","Dic"], index=5)

    @st.cache_data(show_spinner="Cargando Seguimiento Nominal...", ttl=600)
    def load_seg_nominal_for_view(select_mes):
        sn_month = {
            "Jun": "11nk2Z1DmVKaXthy_TRsYK8wYzQ_BW8sdcRSXGxuq4Zo",
            "Jul": "11nk2Z1DmVKaXthy_TRsYK8wYzQ_BW8sdcRSXGxuq4Zo",
            "Ago": "1-jE3mNODE9UXj3dN9G9ae5WLl_Kyow96gWxcB0pOAxo",
            "Set": "1-jE3mNODE9UXj3dN9G9ae5WLl_Kyow96gWxcB0pOAxo",
            "Oct": "1VUARWulWk039rov4fsyM8NJRNd-zWRHyKYhG9tHBefI",
            "Nov": "1XgsYbeYeX9nwyz7jj2PHBYiGcxZcXb9HIiw757XLCcQ",
            "Dic": "1_xhmyCYQAUM6BjDv2mCl5vZ6xeY0lkxVHXgpESDgCFE",
        }
        return read_and_concatenate_sheets_optimized(
            key_sheet=sn_month[select_mes],
            sheet_names=[
                "ARANJUEZ","CLUB DE LEONES","EL BOSQUE","LOS GRANADOS \"SAGRADO CORAZON\"",
                "CENTRO DE SALUD LA UNION","HOSPITAL DE ESPECIALIDADES BASI",
                "LIBERTAD","LOS JARDINES","PESQUEDA III","SAN MARTIN DE PORRES"
            ],
            add_sheet_column=True
        )

    seg_nominal_df = load_seg_nominal_for_view(select_mes)
    if seg_nominal_df is None or seg_nominal_df.empty:
        st.error("No se pudo cargar el Seguimiento Nominal. Revisa las hojas o intenta nuevamente.")
        st.stop()

    # Asegurar columna Actividad a Realizar
    if "Actividad a Realizar" not in seg_nominal_df.columns:
        seg_nominal_df["Actividad a Realizar"] = "Sin rango específico"

    # Limpieza y estandarización de columnas de hemoglobina
    def safe_float_conversion(value):
        try:
            if "/" in str(value):
                return 0.0
            return float(value)
        except (ValueError, TypeError):
            return 0.0

    for col in ["Resultado de Hemoglobina de 06 MESES", "Resultado de Hemoglobina de 12 MESES"]:
        if col in seg_nominal_df.columns:
            seg_nominal_df[col] = seg_nominal_df[col].astype(str).str.strip().replace({"NO": "0", "NO CORRESPONDE": "0", "": "0"})
            seg_nominal_df[col] = seg_nominal_df[col].str.replace("-", ".").str.replace(",", ".")
            seg_nominal_df[col] = seg_nominal_df[col].apply(safe_float_conversion)
        else:
            seg_nominal_df[col] = 0.0

    # Columna de agrupación (EESS)
    eess_col = next((c for c in [
        "Establecimiento de Salud",
        "ESTABLECIMIENTO DE SALUD ",
        "EESS ATENCION",
        "EESS ATENCIONS",
        "EESS NACIMIENTO",
        "sheet_origen"
    ] if c in seg_nominal_df.columns), "sheet_origen")

    # Filtro por actividad
    actividades = ["Todos"] + sorted([a for a in seg_nominal_df["Actividad a Realizar"].dropna().unique().tolist()])
    colfa, colfb = st.columns([3,7])
    with colfa:
        select_act = st.selectbox("Actividad a Realizar", actividades, index=0)
    if select_act != "Todos":
        seg_nominal_df = seg_nominal_df[seg_nominal_df["Actividad a Realizar"] == select_act]

    # Variables derivadas para tamizaje y anemia
    seg_nominal_df["_tam6"] = (seg_nominal_df["Resultado de Hemoglobina de 06 MESES"] > 0).astype(int)
    seg_nominal_df["_ane6"] = ((seg_nominal_df["Resultado de Hemoglobina de 06 MESES"] > 0) & (seg_nominal_df["Resultado de Hemoglobina de 06 MESES"] < 10.5)).astype(int)
    seg_nominal_df["_tam12"] = (seg_nominal_df["Resultado de Hemoglobina de 12 MESES"] > 0).astype(int)
    seg_nominal_df["_ane12"] = ((seg_nominal_df["Resultado de Hemoglobina de 12 MESES"] > 0) & (seg_nominal_df["Resultado de Hemoglobina de 12 MESES"] < 10.5)).astype(int)

    resumen_df = (
        seg_nominal_df.groupby(eess_col)
        .agg(
            Tamizaje_6=("_tam6", "sum"),
            Anemia_6=("_ane6", "sum"),
            Tamizaje_12=("_tam12", "sum"),
            Anemia_12=("_ane12", "sum"),
        )
        .reset_index()
        .rename(columns={eess_col: "Establecimiento de Salud"})
    )

    # Porcentajes de anemia sobre tamizaje (numérico)
    def pct(n, d):
        return round((n / d) * 100, 1) if d and d > 0 else 0.0

    resumen_df["pct_ane6"] = resumen_df.apply(lambda r: pct(r["Anemia_6"], r["Tamizaje_6"]), axis=1)
    resumen_df["pct_ane12"] = resumen_df.apply(lambda r: pct(r["Anemia_12"], r["Tamizaje_12"]), axis=1)

    # Fila total
    tot_tam6 = int(resumen_df["Tamizaje_6"].sum())
    tot_ane6 = int(resumen_df["Anemia_6"].sum())
    tot_tam12 = int(resumen_df["Tamizaje_12"].sum())
    tot_ane12 = int(resumen_df["Anemia_12"].sum())

    total_row = {
        "Establecimiento de Salud": "TOTAL",
        "Tamizaje_6": tot_tam6,
        "Anemia_6": tot_ane6,
        "Tamizaje_12": tot_tam12,
        "Anemia_12": tot_ane12,
        "pct_ane6": pct(tot_ane6, tot_tam6),
        "pct_ane12": pct(tot_ane12, tot_tam12),
    }

    resumen_df = pd.concat([resumen_df, pd.DataFrame([total_row])], ignore_index=True)

    # Gráficos por establecimiento (excluyendo fila TOTAL)
    chart_df = resumen_df[resumen_df["Establecimiento de Salud"] != "TOTAL"].copy()
    colg1, colg2 = st.columns(2)
    with colg1:
        fig_tam6 = px.bar(chart_df, x="Establecimiento de Salud", y="Tamizaje_6", title="Tamizaje 6 Meses", text="Tamizaje_6")
        fig_tam6.update_layout(xaxis_tickangle=-45)
        st.plotly_chart(fig_tam6, use_container_width=True)
    with colg2:
        fig_ane6 = px.bar(chart_df, x="Establecimiento de Salud", y="Anemia_6", title="Niños con Anemia (tamizaje 6M)", text="Anemia_6")
        fig_ane6.update_layout(xaxis_tickangle=-45)
        st.plotly_chart(fig_ane6, use_container_width=True)

    colg3, colg4 = st.columns(2)
    with colg3:
        fig_tam12 = px.bar(chart_df, x="Establecimiento de Salud", y="Tamizaje_12", title="Tamizaje 12 Meses", text="Tamizaje_12")
        fig_tam12.update_layout(xaxis_tickangle=-45)
        st.plotly_chart(fig_tam12, use_container_width=True)
    with colg4:
        fig_ane12 = px.bar(chart_df, x="Establecimiento de Salud", y="Anemia_12", title="Niños con Anemia (tamizaje 12M)", text="Anemia_12")
        fig_ane12.update_layout(xaxis_tickangle=-45)
        st.plotly_chart(fig_ane12, use_container_width=True)

    colg5, colg6 = st.columns(2)
    with colg5:
        fig_pct6 = px.bar(chart_df, x="Establecimiento de Salud", y="pct_ane6", title="% Anemia sobre Tamizaje 6M", text="pct_ane6")
        fig_pct6.update_traces(texttemplate="%{text:.1f}%")
        fig_pct6.update_layout(xaxis_tickangle=-45, yaxis_title="%")
        st.plotly_chart(fig_pct6, use_container_width=True)
    with colg6:
        fig_pct12 = px.bar(chart_df, x="Establecimiento de Salud", y="pct_ane12", title="% Anemia sobre Tamizaje 12M", text="pct_ane12")
        fig_pct12.update_traces(texttemplate="%{text:.1f}%")
        fig_pct12.update_layout(xaxis_tickangle=-45, yaxis_title="%")
        st.plotly_chart(fig_pct12, use_container_width=True)

    # Tabla con porcentajes formateados
    resumen_df["% Anemia 6M"] = resumen_df["pct_ane6"].map(lambda x: f"{x:.1f}%")
    resumen_df["% Anemia 12M"] = resumen_df["pct_ane12"].map(lambda x: f"{x:.1f}%")
    resumen_df = resumen_df.drop(columns=["pct_ane6","pct_ane12"])

    gob = GridOptionsBuilder.from_dataframe(resumen_df)
    gob.configure_default_column(resizable=True, filter=True, sortable=True)
    gob.configure_pagination(paginationPageSize=25)
    grid_options = gob.build()

    AgGrid(
        resumen_df,
        gridOptions=grid_options,
        enable_enterprise_modules=False,
        theme='balham',
        update_mode='NO_UPDATE',
        fit_columns_on_grid_load=True,
        height=500,
        key="grid_nominal_tamizajes",
    )

def eliminar_duplicados_col(texto):
        # Maneja None/NaN y cualquier tipo convirtiendo a string.
        if pd.isna(texto):
            return "-"
        s = str(texto)
        col = [c for c in s.split(" - ") if c.strip() != ""]
        if not col:
            return "-"
        col_unicos = sorted(set(col), key=col.index)
        return " - ".join(col_unicos)

def tomar_ultimo_elemento(texto):
    col = [c for c in texto.split(" - ") if c.strip() != ""]
    return col[-1] if col else ""

def visitas_ninos_dashboard():
        
                styles(2)
                
                # Cache the data loading operations
                #@st.cache_data(show_spinner="Cargando datos...",ttl=600)  # Cache for 5 minutes
                def load_base_data():
                    
                    actvd_df = fetch_vd_childs()
                    carga_df = fetch_carga_childs()
                    
                    padron_df = fetch_padron()
                    datos_ninos_df = pd.read_parquet('datos_niños.parquet', engine='pyarrow')
                    return actvd_df, carga_df, padron_df, datos_ninos_df
                
                @st.cache_data(show_spinner="Cargando datos...")  # ,ttl=600
                def load_seg_nominal_data(select_mes):
                    sn_month = {
                            "Jun": "11nk2Z1DmVKaXthy_TRsYK8wYzQ_BW8sdcRSXGxuq4Zo",
                            "Jul": "11nk2Z1DmVKaXthy_TRsYK8wYzQ_BW8sdcRSXGxuq4Zo",
                            "Ago": "1-jE3mNODE9UXj3dN9G9ae5WLl_Kyow96gWxcB0pOAxo",
                            "Set": "1XgsYbeYeX9nwyz7jj2PHBYiGcxZcXb9HIiw757XLCcQ",
                            "Oct": "1VUARWulWk039rov4fsyM8NJRNd-zWRHyKYhG9tHBefI",
                            "Nov": "1XgsYbeYeX9nwyz7jj2PHBYiGcxZcXb9HIiw757XLCcQ",
                            "Dic": "1XgsYbeYeX9nwyz7jj2PHBYiGcxZcXb9HIiw757XLCcQ",
                            "Ene": "1XgsYbeYeX9nwyz7jj2PHBYiGcxZcXb9HIiw757XLCcQ",
                            "Feb": "1XgsYbeYeX9nwyz7jj2PHBYiGcxZcXb9HIiw757XLCcQ",
                    }
                    print(sn_month[select_mes])
                    return read_and_concatenate_sheets_optimized(
                        
                        key_sheet=sn_month[select_mes],
                        sheet_names=[
                            "ARANJUEZ","CLUB DE LEONES","EL BOSQUE",'LOS GRANADOS "SAGRADO CORAZON"',
                            "CENTRO DE SALUD LA UNION","HOSPITAL DE ESPECIALIDADES BASI",
                            "LIBERTAD","LOS JARDINES","PESQUEDA III","SAN MARTIN DE PORRES"
                        ],
                        add_sheet_column=True  # Añade columna 'sheet_origen'
                    )
            
            # Load cached data
            #try:
                actvd_df, carga_df, padron_df, datos_ninos_df = load_base_data()
                
                #hbdff = load_hb_data()
                #supledf = load_suplementacion_data()
                
                # Continue with the rest of the processing...
                fecha_update = str(carga_df["update"].unique()[0])[:-7]
                fecha_actual = str(datetime.now().strftime("%Y-%m-%d %H:%M"))
                
                
                #fecha_update = dt.strftime("%Y-%m-%d-%H-%M")
                #list_mes = [mes_short(x) for x in sorted(list(carga_df["Mes"].unique()))]
                
                col_head1, col_head2, col_head3, col_head4 = st.columns([3,3,1,2])
                with col_head1:
                    st.title("Visitas a Niños")
                with col_head2:
                    st.subheader(f"Act: {fecha_update[:-6]}", divider=True)
                with col_head3:
                    select_year  = st.selectbox("Año:", ["2026"], key="select1")
                    
                with col_head4:
                    select_mes  = st.selectbox("Mes:",['Feb'] , key="select2",index=0)
                #with col_head5:
                #
                #     select_rango  = st.selectbox("Rango:",['1-5 meses','6-11 meses'] , key="select7",index=None)
                #st.dataframe(datos_ninos_df)
                datos_ninos_df = datos_ninos_df[datos_ninos_df["Periodo"]==f"{select_year}-{select_mes}"]
                #st.write(f"{select_year}-{select_mes}")
                
                datos_ninos_df = fix_data_childs(datos_ninos_df)

                seg_nominal_df = load_seg_nominal_data(select_mes)
                #st.write(seg_nominal_df.shape)
                #st.dataframe(seg_nominal_df)
                # Guardar contra fallos de carga/concatenación
                if seg_nominal_df is None:
                    st.error("No se pudo cargar el Seguimiento Nominal. Revisa las hojas o intenta nuevamente.")
                    st.stop()
                #st.dataframe(seg_nominal_df)
                #seg_nominal_test = seg_nominal_df.copy()
                
                # Procesar columna de 06 MESES
                seg_nominal_df["Resultado de Hemoglobina de 06 MESES"] = seg_nominal_df["Resultado de Hemoglobina de 06 MESES"].astype(str).str.strip().fillna("0")
                seg_nominal_df["Resultado de Hemoglobina de 06 MESES"] = seg_nominal_df["Resultado de Hemoglobina de 06 MESES"].astype(str).str.strip()
                seg_nominal_df["Resultado de Hemoglobina de 06 MESES"] = seg_nominal_df["Resultado de Hemoglobina de 06 MESES"].astype(str).replace("NO", "0")
                seg_nominal_df["Resultado de Hemoglobina de 06 MESES"] = seg_nominal_df["Resultado de Hemoglobina de 06 MESES"].astype(str).replace("NO CORRESPONDE", "0")
                seg_nominal_df["Resultado de Hemoglobina de 06 MESES"] = seg_nominal_df["Resultado de Hemoglobina de 06 MESES"].astype(str).replace("", "0")
                seg_nominal_df["Resultado de Hemoglobina de 06 MESES"] = seg_nominal_df["Resultado de Hemoglobina de 06 MESES"].astype(str).str.replace("-", ".")
                seg_nominal_df["Resultado de Hemoglobina de 06 MESES"] = seg_nominal_df["Resultado de Hemoglobina de 06 MESES"].astype(str).str.replace(",", ".")
                
                # Convertir a float con manejo de errores para formatos de fecha
                def safe_float_conversion(value):
                    try:
                        # Si contiene "/" es probablemente una fecha, convertir a 0
                        if "/" in str(value):
                            return 0.0
                        return float(value)
                    except (ValueError, TypeError):
                        return 0.0
                
                seg_nominal_df["Resultado de Hemoglobina de 06 MESES"] = seg_nominal_df["Resultado de Hemoglobina de 06 MESES"].apply(safe_float_conversion)
                
                # Procesar columna de 12 MESES
                seg_nominal_df["Resultado de Hemoglobina de 12 MESES"] = seg_nominal_df["Resultado de Hemoglobina de 12 MESES"].astype(str).str.strip().fillna("0")
                seg_nominal_df["Resultado de Hemoglobina de 12 MESES"] = seg_nominal_df["Resultado de Hemoglobina de 12 MESES"].astype(str).str.strip()
                seg_nominal_df["Resultado de Hemoglobina de 12 MESES"] = seg_nominal_df["Resultado de Hemoglobina de 12 MESES"].astype(str).replace("NO", "0")
                seg_nominal_df["Resultado de Hemoglobina de 12 MESES"] = seg_nominal_df["Resultado de Hemoglobina de 12 MESES"].astype(str).replace("NO CORRESPONDE", "0")
                seg_nominal_df["Resultado de Hemoglobina de 12 MESES"] = seg_nominal_df["Resultado de Hemoglobina de 12 MESES"].astype(str).replace("", "0")
                seg_nominal_df["Resultado de Hemoglobina de 12 MESES"] = seg_nominal_df["Resultado de Hemoglobina de 12 MESES"].astype(str).str.replace("-", ".")
                seg_nominal_df["Resultado de Hemoglobina de 12 MESES"] = seg_nominal_df["Resultado de Hemoglobina de 12 MESES"].astype(str).str.replace(",", ".")
                seg_nominal_df["Resultado de Hemoglobina de 12 MESES"] = seg_nominal_df["Resultado de Hemoglobina de 12 MESES"].apply(safe_float_conversion)
               
                seg_nominal_df["Resultado de Hemoglobina de 09 MESES"] = seg_nominal_df["Resultado de Hemoglobina de 09 MESES"].astype(str).str.strip().fillna("0")
                seg_nominal_df["Resultado de Hemoglobina de 09 MESES"] = seg_nominal_df["Resultado de Hemoglobina de 09 MESES"].astype(str).str.strip()
                seg_nominal_df["Resultado de Hemoglobina de 09 MESES"] = seg_nominal_df["Resultado de Hemoglobina de 09 MESES"].astype(str).replace("NO", "0")
                seg_nominal_df["Resultado de Hemoglobina de 09 MESES"] = seg_nominal_df["Resultado de Hemoglobina de 09 MESES"].astype(str).replace("NO CORRESPONDE", "0")
                seg_nominal_df["Resultado de Hemoglobina de 09 MESES"] = seg_nominal_df["Resultado de Hemoglobina de 09 MESES"].astype(str).replace("", "0")
                seg_nominal_df["Resultado de Hemoglobina de 09 MESES"] = seg_nominal_df["Resultado de Hemoglobina de 09 MESES"].astype(str).str.replace("-", ".")
                seg_nominal_df["Resultado de Hemoglobina de 09 MESES"] = seg_nominal_df["Resultado de Hemoglobina de 09 MESES"].astype(str).str.replace(",", ".")
                seg_nominal_df["Resultado de Hemoglobina de 09 MESES"] = seg_nominal_df["Resultado de Hemoglobina de 09 MESES"].apply(safe_float_conversion)
                
                




                seg_nominal_df = seg_nominal_df[["Número de Documento del niño","TIPO DE SEGURO","¿Es prematuro?","Tipo de SUPLEMENTO","¿Fue parte de una Sesion demostrativa?","Resultado de Hemoglobina de 06 MESES","Resultado de Hemoglobina de 09 MESES","Resultado de Hemoglobina de 12 MESES"]]
                seg_nominal_df["TIPO DE SEGURO"] = seg_nominal_df["TIPO DE SEGURO"].replace("", "NINGUNO")
                seg_nominal_df["¿Es prematuro?"] = seg_nominal_df["¿Es prematuro?"].replace("", "NO")
                seg_nominal_df["Tipo de SUPLEMENTO"] = seg_nominal_df["Tipo de SUPLEMENTO"].replace("", "NO ESPECIFICADO")
                seg_nominal_df["¿Fue parte de una Sesion demostrativa?"] = seg_nominal_df["¿Fue parte de una Sesion demostrativa?"].replace("", "NO")
                seg_nominal_df["¿Fue parte de una Sesion demostrativa?"] = seg_nominal_df["¿Fue parte de una Sesion demostrativa?"].str.strip()
                seg_nominal_df = seg_nominal_df.rename(columns={"Número de Documento del niño":"Número de Documento"})
                seg_nominal_df["Tipo de SUPLEMENTO"] = seg_nominal_df["Tipo de SUPLEMENTO"].str.strip()
                seg_nominal_df["¿Fue parte de una Sesion demostrativa?"] = seg_nominal_df["¿Fue parte de una Sesion demostrativa?"].replace("Si", "SI")
                seg_nominal_df["¿Fue parte de una Sesion demostrativa?"] = seg_nominal_df["¿Fue parte de una Sesion demostrativa?"].fillna("NO")
                
                #st.dataframe(seg_nominal_test)
                #st.dataframe(datos_ninos_df)
                carga_filt_df = carga_df[(carga_df['Año']==int(select_year))&(carga_df['Mes']==int(mestext_short(select_mes)))]
                actvd_filt_df = actvd_df[(actvd_df['Año']==select_year)&(actvd_df['Mes']==select_mes)]
                
                

                niños_unicos_vd = childs_unicos_visitados(actvd_filt_df,'Número de Documento de Niño',"ALL CHILDS W DUPLICADOS")

                ## CALCULAR INDICADORES
                num_carga = carga_filt_df.shape[0]
                num_visitas_programadas = carga_filt_df["Total de visitas completas para la edad"].sum()
                num_visitas = carga_filt_df["Total de Intervenciones"].sum()
                num_child_vd = niños_unicos_vd.shape[0]
                num_vd_movil = carga_filt_df["Total de VD presencial Válidas MOVIL"].sum()
                porcentaje_child_visita = f"{round((num_child_vd/num_carga)*100,2)}%"

                ###
                

                ###
                actvd_filt_df = actvd_filt_df.groupby(['Número de Documento de Niño','Actores Sociales','Etapa'])[['Año']].count().reset_index()
                actvd_filt_df = actvd_filt_df.sort_values(by='Etapa', key=lambda x: x.isin(['No Encontrado', 'Rechazado']))
                actvd_filt_df = actvd_filt_df.drop_duplicates(subset='Número de Documento de Niño', keep='first')
                actvd_filt_df.columns = ["Doc_Ultimo_Mes","Actor Social Ultimo Mes","Estado_Visita_Ult","count"]
                actvd_filt_df = actvd_filt_df[["Doc_Ultimo_Mes","Actor Social Ultimo Mes","Estado_Visita_Ult"]]
                actvd_filt_df["Doc_Ultimo_Mes"] = actvd_filt_df["Doc_Ultimo_Mes"].astype(str)
                print(padron_df.columns)
                dataframe_pn = padron_df[COLUMNS_PADRON_C1_VD]  
                join_df = pd.merge(carga_filt_df, dataframe_pn, left_on='Número de Documento del niño', right_on='Documento', how='left')
                
                
                join_df['Prioridad'] = join_df['Tipo de Documento'].map(PRIORIDAD_TIPO_DOCUMENTO)
                
                ###
                new_col_estado_padron = join_df.groupby(['Número de Documento del niño'])[['Prioridad']].count().reset_index()
                new_col_estado_padron["Prioridad"] = new_col_estado_padron["Prioridad"].replace({0: "NO EXISTE REGISTRO",1: "REGISTRO UNICO", 2: "REGISTRO DUPLICADO", 3: "REGISTRO TRIPLICADO"})
                new_col_estado_padron.columns = ["Número de Documento del niño","Estado Padrón Nominal"]
                new_col_estado_padron["Número de Documento del niño"] = new_col_estado_padron["Número de Documento del niño"].astype(int).astype(str)

                ###
                join_df = join_df.sort_values(by=['Número de Documento del niño', 'Prioridad'])
                join_df = join_df.drop_duplicates(subset='Número de Documento del niño', keep='first')
                join_df = join_df.drop(columns=['Prioridad'])
                join_df['Número de Documento del niño'] = join_df['Número de Documento del niño'].astype(int).astype(str)
                join_df = pd.merge(join_df, new_col_estado_padron, left_on='Número de Documento del niño', right_on='Número de Documento del niño', how='left')
                ##
                dataframe_ = pd.merge(join_df, actvd_filt_df, left_on='Número de Documento del niño', right_on='Doc_Ultimo_Mes', how='left')
                
                #datos_ninos_df["Documento_c1"] = datos_ninos_df["Documento_c1"].astype(int).astype(str)
                
                #dataframe_["Número de Documento del niño"] = dataframe_["Número de Documento del niño"].astype(int).astype(str)
                
                del join_df

                dataframe_ = pd.merge(dataframe_, datos_ninos_df, left_on='Número de Documento del niño', right_on='Documento_c1', how='left')
                dataframe_['DATOS NIÑO PADRON'] = dataframe_['DATOS NIÑO PADRON'].fillna(dataframe_['Niño'])
                dataframe_['Tipo de Documento'] = dataframe_['Tipo de Documento'].fillna(dataframe_['Tipo de Documento del niño'])
                dataframe_['MENOR VISITADO'] = dataframe_['MENOR VISITADO'].fillna("SIN MODIFICACION EN EL PADRON")
                dataframe_['TIPO DE DOCUMENTO DE LA MADRE'] = dataframe_['TIPO DE DOCUMENTO DE LA MADRE'].fillna("MADRE DE GOLONDRINO")
                dataframe_['EESS NACIMIENTO'] = dataframe_['EESS NACIMIENTO'].fillna("No Especificado")
                dataframe_['EESS'] = dataframe_['EESS'].fillna("No Especificado")
                dataframe_['FRECUENCIA DE ATENCION'] = dataframe_['FRECUENCIA DE ATENCION'].fillna("No Especificado")
                dataframe_['EESS ADSCRIPCIÓN'] = dataframe_['EESS ADSCRIPCIÓN'].fillna("No Especificado")
                dataframe_['Tipo_file'] = dataframe_['Tipo_file'].fillna("En Otro Padrón Nominal")
                dataframe_['ENTIDAD'] = dataframe_['ENTIDAD'].fillna("En Otro Padrón Nominal")
                dataframe_['USUARIO QUE MODIFICA'] = dataframe_['USUARIO QUE MODIFICA'].fillna(f"Usuario no definido")
                
                del datos_ninos_df
                
                dataframe_ = dataframe_[COL_ORDER_VD_CHILD_C1]
                dataframe_.columns = COL_ORDER_VD_CHILD_C1_2
                dataframe_['Estado Visitas'] =  dataframe_.apply(lambda x: estado_visitas_completas(x['N° Visitas Completas'], x['Total de VD presenciales Válidas'],x['Estado Niño']),axis=1)
                dataframe_['Estado Niño'] = dataframe_['Estado Niño'].fillna(f"Sin Visita ({select_mes})")
                dataframe_['Edad'] = dataframe_['Fecha de Nacimiento'].apply(calcular_edad)

                dataframe_['Edad Dias'] = dataframe_['Fecha de Nacimiento'].apply(lambda x: (datetime.now() - x).days)
            
                primer_dia_mes = datetime(int(select_year), int(mestext_short(select_mes)), 1)
                if int(mestext_short(select_mes)) == 12:
                    ultimo_dia_mes = datetime(int(select_year) + 1, 1, 1) - pd.Timedelta(days=1)
                else:
                    ultimo_dia_mes = datetime(int(select_year), int(mestext_short(select_mes)) + 1, 1) - pd.Timedelta(days=1)

                dataframe_['Edad en días (primer día del mes)'] = dataframe_['Fecha de Nacimiento'].apply(lambda x: (primer_dia_mes - x).days)
                dataframe_['Edad en días (último día del mes)'] = dataframe_['Fecha de Nacimiento'].apply(lambda x: (ultimo_dia_mes - x).days)


                dataframe_['Niños 120-149 días en mes'] = dataframe_.apply(
                    lambda row: "SI" if (
                        (row['Edad en días (primer día del mes)'] <= 149 and row['Edad en días (último día del mes)'] >= 120)
                    ) else "NO", axis=1
                )

                dataframe_['Niños 180-209 días en mes'] = dataframe_.apply(
                    lambda row: "SI" if (
                        (row['Edad en días (primer día del mes)'] <= 209 and row['Edad en días (último día del mes)'] >= 180)
                    ) else "NO", axis=1
                )

                dataframe_['Niños 270-299 días en mes'] = dataframe_.apply(
                    lambda row: "SI" if (
                        (row['Edad en días (primer día del mes)'] <= 299 and row['Edad en días (último día del mes)'] >= 270)
                    ) else "NO", axis=1
                )

                dataframe_['Niños 360-389 días en mes'] = dataframe_.apply(
                    lambda row: "SI" if (
                        (row['Edad en días (primer día del mes)'] <= 389 and row['Edad en días (último día del mes)'] >= 360)
                    ) else "NO", axis=1
                )
                dataframe_['Rango de Días Activo'] = dataframe_.apply(combinar_rangos_dias, axis=1)
                
                # Crear columna FECHA CUMPLE
                def calcular_fecha_cumple(row):
                    rango = row['Rango de Días Activo']
                    fecha_nac = row['Fecha de Nacimiento']
                    
                    if rango == "Rango de días 120-149 días":
                        return fecha_nac + pd.Timedelta(days=120)
                    elif rango == "Rango de días 180-209 días":
                        return fecha_nac + pd.Timedelta(days=180)
                    elif rango == "Rango de días 270-299 días":
                        return fecha_nac + pd.Timedelta(days=270)
                    elif rango == "Rango de días 360-389 días":
                        return fecha_nac + pd.Timedelta(days=360)
                    else:
                        return pd.NaT  # Para casos sin rango específico
                
                dataframe_['FECHA CUMPLE'] = dataframe_.apply(calcular_fecha_cumple, axis=1)
                
                dataframe_ = pd.merge(dataframe_, seg_nominal_df, on='Número de Documento', how='left')
                
                # Use the cached hb data
                #dataframe_ = pd.merge(dataframe_,hbdff,on="Número de Documento",how="left")

                # Aplicar la función para crear las nuevas columnas
                #dataframe_[['Ultima Fecha tamizaje', 'Ultima HB']] = dataframe_['Resultados'].apply(
                #    lambda x: pd.Series(extraer_ultimo_resultado(x))
                #)
                # Crear la columna ESTADO TAMIZAJE
                #dataframe_['Edad Ultimo Tamizaje'] = dataframe_.apply(
                #    lambda row: calcular_edad_diagnostico(row['Fecha de Nacimiento'], row['Ultima Fecha tamizaje']), 
            #     axis=1
            # )
                #dataframe_['ESTADO TAMIZAJE'] = dataframe_['Ultima HB'].apply(determinar_estado_tamizaje)

                # Use the cached suplementacion data
                #if 'Fecha_Diagnostico' in supledf.columns and 'FECHA_NAC' in supledf.columns:
            
                #    supledf['Edad Diagnóstico'] = supledf.apply(
                #        lambda row: calcular_edad_diagnostico(row['FECHA_NAC'], row['Fecha_Diagnostico']), 
                #        axis=1
                #    )
                    
                #    supledf['Edad Diagnóstico (días)'] = supledf.apply(
                #        lambda row: calcular_edad_diagnostico_dias(row['FECHA_NAC'], row['Fecha_Diagnostico']), 
                #        axis=1
                #    )
                #supledf["ACTIVIDAD"] = supledf["ACTIVIDAD"].str[4:]
                #supledf["ACTIVIDAD"] = supledf["ACTIVIDAD"].str.strip()
                #supledf = supledf[supledf["MICRORED"].notna()]
                #print(supledf["DIAGNOSTICO"].unique())
                # Crear columnas pivote para cada diagnóstico único
                #diagnosticos_unicos = supledf["DIAGNOSTICO"].unique()
                # Agrupar por Documento y obtener la fecha más reciente para cada diagnóstico
                #supledf_pivot = supledf.groupby(['Documento', 'DIAGNOSTICO'])['Fecha_Diagnostico'].max().reset_index()
                # Hacer pivot de los diagnósticos
                #supledf_pivot = supledf_pivot.pivot(index='Documento', columns='DIAGNOSTICO', values='Fecha_Diagnostico').reset_index()
                # Aplicar la función para crear la columna ESTADO HB NIÑO
                #supledf_pivot['ESTADO HB NIÑO'] = supledf_pivot.apply(determinar_estado_hb, axis=1)
                #supledf_pivot = supledf_pivot.rename(columns={"Documento":"Número de Documento"})
                #supledf_pivot["Número de Documento"] = supledf_pivot["Número de Documento"].astype(str)
                
                #dataframe_ = pd.merge(dataframe_, supledf_pivot, on='Número de Documento', how='left')
                #dataframe_["ESTADO HB NIÑO"] = dataframe_["ESTADO HB NIÑO"].fillna("SIN DIAGNOSTICO")
                dataframe_["Actividad a Realizar"] = dataframe_["Rango de Días Activo"].replace(
                    {"Rango de días 120-149 días":"Suplementación Preventiva",
                    "Rango de días 180-209 días":"Tamizaje de 6 m.",
                    "Rango de días 270-299 días":"Tamizaje de Control",
                    "Rango de días 360-389 días":"Tamizaje de 12 m.",
                    "Sin rango específico":"Sin rango específico"
                })
                #st.dataframe(dataframe_)
                metric_col = st.columns(7)
                metric_col[0].metric("Niños Cargados",num_carga,f"Con Visita {num_child_vd}({num_carga-num_child_vd})",border=True)
                metric_col[1].metric("Total de Visitas",num_visitas,f"VD Programadas: {num_visitas_programadas}",border=True)
                
                st.session_state["dataframe_childs"] = dataframe_.copy()

                #########################################################   
                tab1= st.tabs(["Seguimiento Visitas Georreferenciadas"])#,"Seguimiento Indicadores Anemia"
                with tab1[0]:
                    st.subheader("📱Seguimiento Visitas Georreferenciadas")
                    vd_geo_percent_df = dataframe_[(dataframe_["Estado Visitas"]=="Visitas Completas")]#&(dataframe_["Celular Madre"]!=0)
                    num_ninos_result = vd_geo_percent_df.shape[0]
                    total_vd_movil_completas = vd_geo_percent_df["Total de VD presencial Válidas MOVIL"].sum()
                    #total_meta_vd = round(num_visitas_validas*0.75)
                    total_meta_vd = round(num_visitas_programadas*PORCENTAJE_GEOS_VD)
                    #num_visitas_programadas
                    total_faltante_vd_meta = total_meta_vd-total_vd_movil_completas
                    percent_vd_movil_validate = safe_percent(total_vd_movil_completas, num_visitas_programadas)
                    percent_vd_completas_movil = safe_percent(total_vd_movil_completas, num_vd_movil)
                    
                    ###
                    

                    #con_visita_cel = dataframe_[(dataframe_["Total de Intervenciones"]!=0)]
                    #num_con_visita_cel = con_visita_cel.shape[0]
                    con_celular = (dataframe_["Celular Madre"]!=0).sum()
                    percent_reg_tel = safe_percent(con_celular, num_carga)
                    percent_total_vd_12 = safe_percent(num_ninos_result, num_carga)
                    num_niños_padron = (dataframe_["CUMPLE INDICADOR"]=="Cumple").sum()
                    num_ninos_cumple = round((num_niños_padron/num_carga),3)*100
                    ########################################################
                    metric_col[2].metric("Visitas Movil",num_vd_movil,f"VD Completas:{total_vd_movil_completas}({percent_vd_completas_movil}%)",border=True)
                    metric_col[6].metric("% Niños Actualizados",num_ninos_cumple,f"N° Act: {num_niños_padron}",border=True)
                    metric_col[3].metric("% VD Georreferenciadas",f"{percent_vd_movil_validate}%",f"VD Faltantes {total_faltante_vd_meta}",border=True)
                    metric_col[4].metric("% Registros Telefonicos",f"{percent_reg_tel}%",f"Sin celular : {num_carga-con_celular}",border=True)
                    metric_col[5].metric("% Niños Oportunos y Completos",f"{percent_total_vd_12}%",f"Positivos:{num_ninos_result}",border=True)
                    
                    #########################################################################################################################
                    dataframe_efec = dataframe_[dataframe_["Estado Niño"].isin([
                        "Visita Domiciliaria (6 a 12 Meses)", "Visita Domiciliaria (1 a 5 meses)"
                    ])]

                    # Groupby y renombrado
                    child_programados_df = dataframe_.groupby("Establecimiento de Salud")["N° Visitas Completas"].count().reset_index(name="Niños Programados")
                    vd_programadas_df = dataframe_.groupby("Establecimiento de Salud")["N° Visitas Completas"].sum().reset_index(name="Visitas Programadas")
                    childs_encontrados_df = dataframe_efec.groupby("Establecimiento de Salud")["Estado Niño"].count().reset_index(name="Niños Encontrados")
                    vd_movil_df = dataframe_efec.groupby("Establecimiento de Salud")["Total de VD presencial Válidas MOVIL"].sum().reset_index(name="Visitas Realizadas MOVIL GEO")

                    # Merge de todos los datos
                    dfs = [child_programados_df, vd_programadas_df, childs_encontrados_df, vd_movil_df]
                    from functools import reduce
                    proyectado_dff = reduce(lambda left, right: pd.merge(left, right, on="Establecimiento de Salud", how="left"), dfs)

                    # Rellenar nulos
                    proyectado_dff[["Niños Encontrados", "Visitas Realizadas MOVIL GEO"]] = proyectado_dff[["Niños Encontrados", "Visitas Realizadas MOVIL GEO"]].fillna(0)

                    # Cálculos vectorizados
                    proyectado_dff["Valla Visitas GEO 85%"] = (proyectado_dff["Visitas Programadas"] * PORCENTAJE_GEOS_VD).round()
                    proyectado_dff["Visitas Faltantes Valla GEO"] = proyectado_dff["Valla Visitas GEO 85%"] - proyectado_dff["Visitas Realizadas MOVIL GEO"]

                    # Visitas proyectadas y tolerancias
                    visitas_for_completar_df = dataframe_efec[dataframe_efec["Estado Visitas"] != "Visitas Completas"].copy()
                    visitas_for_completar_df["Visitas GEO Proyectadas"] = visitas_for_completar_df["N° Visitas Completas"] - visitas_for_completar_df["Total de VD presenciales Válidas"]
                    vd_completar_df = visitas_for_completar_df.groupby("Establecimiento de Salud")["Visitas GEO Proyectadas"].sum().reset_index()

                    proyectado_dff = proyectado_dff.merge(vd_completar_df, on="Establecimiento de Salud", how="left")
                    proyectado_dff["Visitas GEO Proyectadas"] = proyectado_dff["Visitas GEO Proyectadas"].fillna(0)
                    #proyectado_dff["Tolerancia Visitas WEB"] = proyectado_dff["Visitas GEO Proyectadas"] - proyectado_dff["Visitas Faltantes Valla GEO"]
                    #proyectado_dff["Tolerancia Niños WEB"] = (proyectado_dff["Tolerancia Visitas WEB"] / 3).round()

                    # Fila de totales
                    total_row = pd.DataFrame({
                        "Establecimiento de Salud": ["TOTAL"],
                        "Niños Programados": [proyectado_dff["Niños Programados"].sum()],
                        "Visitas Programadas": [proyectado_dff["Visitas Programadas"].sum()],
                        "Niños Encontrados": [proyectado_dff["Niños Encontrados"].sum()],
                        "Visitas Realizadas MOVIL GEO": [proyectado_dff["Visitas Realizadas MOVIL GEO"].sum()],
                        "Valla Visitas GEO 85%": [proyectado_dff["Valla Visitas GEO 85%"].sum()],
                        "Visitas Faltantes Valla GEO": [proyectado_dff["Visitas Faltantes Valla GEO"].sum()],
                        "Visitas GEO Proyectadas": [proyectado_dff["Visitas GEO Proyectadas"].sum()],
                        #"Tolerancia Visitas WEB": [proyectado_dff["Tolerancia Visitas WEB"].sum()],
                        #"Tolerancia Niños WEB": [proyectado_dff["Tolerancia Niños WEB"].sum()],
                    })

                    proyectado_dff = pd.concat([proyectado_dff, total_row], ignore_index=True)

                    # Estado y porcentaje
                    proyectado_dff['Estado'] = proyectado_dff.apply(
                        lambda x: estado_proyectado(x['Visitas Faltantes Valla GEO'], x['Visitas GEO Proyectadas']), axis=1
                    )
                    proyectado_dff["% Actual GEO"] = ((proyectado_dff["Visitas Realizadas MOVIL GEO"] / proyectado_dff["Visitas Programadas"]) * 100).round(1).astype(str) + "%"

                    # Selección y orden de columnas
                    cols = [
                        'Establecimiento de Salud', 'Niños Programados', 'Visitas Programadas',
                        'Niños Encontrados', 'Visitas Realizadas MOVIL GEO', '% Actual GEO',
                        'Valla Visitas GEO 85%', 'Visitas Faltantes Valla GEO', 'Visitas GEO Proyectadas',
                        'Estado'
                    ]
                    proyectado_dff = proyectado_dff[cols]

                    # Ordenar por 'Niños Programados', dejando TOTAL al final
                    proyectado_dff_no_total = proyectado_dff[proyectado_dff["Establecimiento de Salud"] != "TOTAL"]
                    total_row = proyectado_dff[proyectado_dff["Establecimiento de Salud"] == "TOTAL"]
                    proyectado_dff_no_total = proyectado_dff_no_total.sort_values("Niños Programados", ascending=False)
                    proyectado_dff = pd.concat([proyectado_dff_no_total, total_row], ignore_index=True)

                    
                    gb = GridOptionsBuilder.from_dataframe(proyectado_dff)
                    gb.configure_default_column(cellStyle={'fontSize': '17px'}) 
                    gb.configure_column("Establecimiento de Salud", width=340)
                    # Formato condicional para la columna Estado
                    
                    grid_options = gb.build()
                    #grid_options['getRowStyle'] = row_style

                    grid_response = AgGrid(proyectado_dff, # Dataframe a mostrar
                                            gridOptions=grid_options,
                                            enable_enterprise_modules=False,
                                            #theme='balham',  # Cambiar tema si se desea ('streamlit', 'light', 'dark', 'alpine', etc.)
                                            update_mode='MODEL_CHANGED',
                                            fit_columns_on_grid_load=True,
                                            height=370,
                                            key="grid_geo"
                                        )
                """
                with tab2:
                    
                    @st.fragment
                    def render_anemia_indicators():
                        col1, col2, = st.columns([2,2])
                        with col1:
                            st.subheader("📱Seguimiento Indicadores Anemia")
                        with col2:
                            selection_rango = st.segmented_control(
                                "Rangos de Edad",
                                options=["120-149 días","180-209 días","270-299 días","360-389 días","6-12 meses"],
                                default="120-149 días",
                                #format_func=lambda option: option_map[option],
                                selection_mode="single",
                            )
                        #with col3:
                        #    eess = st.selectbox("Establecimiento de Salud",options = dataframe_["Establecimiento de Salud"].unique().tolist(),index=None,placeholder="Todos")

                        if selection_rango == "120-149 días":
                            rg_df = dataframe_[dataframe_["Rango de Días Activo"]=="Rango de días 120-149 días"]
                        elif selection_rango == "180-209 días":
                            rg_df = dataframe_[dataframe_["Rango de Días Activo"]=="Rango de días 180-209 días"]
                        elif selection_rango == "270-299 días":
                            rg_df = dataframe_[dataframe_["Rango de Días Activo"]=="Rango de días 270-299 días"]
                            # Asegura columna de Hb 09 meses si falta para evitar KeyError
                            if "Resultado de Hemoglobina de 09 MESES" not in rg_df.columns:
                                rg_df["Resultado de Hemoglobina de 09 MESES"] = 0
                        elif selection_rango == "360-389 días":
                            rg_df = dataframe_[dataframe_["Rango de Días Activo"]=="Rango de días 360-389 días"]
                        elif selection_rango == "6-12 meses":
                            rg_df = dataframe_[dataframe_["Rango de Edad"]=="6-12 meses"]
                        rg_dff = rg_df.copy()
                        if selection_rango == "270-299 días":
                            rg_df = rg_df.groupby(["Establecimiento de Salud"]).agg(
                                Niños_Programados=("Número de Documento", "count"),
                                Niños_Encontrados=("Estado Niño", lambda x: (x.isin(["Visita Domiciliaria (6 a 12 Meses)", "Visita Domiciliaria (1 a 5 meses)"])).sum()),
                                #Niños_Consecutivos=("¿Es consecutivo?", lambda x: (x.isin(["Consecutivo"])).sum()),
                                Con_Suplementacion=("Tipo de SUPLEMENTO", lambda x: (x.isin([
                                'MMN', 'MICRONUTRIENTES', 'GOTAS', 'FERRIMAX',
                                    'SULFATO FERROSO', 'MN', 'JARABE',
                                    'FERRAMIN FORTE', 'POLIMALTOSADO', 'SI',
                                    'FERRAMIN', '7 GOTAS', '8 GOTAS', '8 GOTITAS DE HIERRO',
                                    'MULTIMICRONUTRIENTES', 'GOTA', 'MALTOFER', 'SULFATOFERROSO',
                                    '1 MICRONUTRIENTE', 'M', 'FERRANIN FORTE',
                                    '1 SOBRECITO DE HIERRO', 'FERANIN', 'HIERRO EN GOTAS',
                                    '9 GOTITAS DE HIERRO', 'HIERRO POLIMALTOSA',
                                    'MULTIVITAMINAS', '7 GOTITAS DE HIERRO',  'FERROCIN',
                                    'GOTAS- SF','gotas', '6 GOTAS',
                                    'HIERRO DE ZINC', 'FERROCIL', 'FERROMIL'
                                ])).sum()),
                                Con_Sesion_Demostrativa=("¿Fue parte de una Sesion demostrativa?", lambda x: (x.isin(["SI"])).sum()),
                                Con_Tamizaje_6_Meses=("Resultado de Hemoglobina de 06 MESES", lambda x: (pd.to_numeric(x, errors='coerce') > 0).sum()),
                                Con_Tamizaje_6_Meses_anemia=("Resultado de Hemoglobina de 06 MESES", lambda x: ((pd.to_numeric(x, errors='coerce') < 10.5) & (pd.to_numeric(x, errors='coerce') > 0)).sum()),
                                
                                
                                Con_Tamizaje_09_Meses=("Resultado de Hemoglobina de 09 MESES", lambda x: (pd.to_numeric(x, errors='coerce') > 0).sum()),
                                Con_Tamizaje_09_Meses_anemia=("Resultado de Hemoglobina de 09 MESES", lambda x: ((pd.to_numeric(x, errors='coerce') < 10.5) & (pd.to_numeric(x, errors='coerce') > 0)).sum()),

                            ).reset_index()
                            rg_df.columns = ["Establecimiento de Salud","Programados", "Encontrados","Con Suplementación","Con Sesion Demostrativa","Con Tamizaje 6 Meses","Con Anemia 6 Meses","Con Tamizaje 09 Meses","Con Anemia 09 Meses"]#,"Sin Anemia (Tamizaje)","Con Anemia (Tamizaje)","Sin Anemia (Diagnostico)","Con Anemia (Diagnostico)"
                            rg_df = rg_df.sort_values("Programados", ascending=False)
                            total_row = pd.DataFrame({
                                "Establecimiento de Salud": ["TOTAL"],
                                "Programados": [rg_df["Programados"].sum()],
                                "Encontrados": [rg_df["Encontrados"].sum()],
                                #"Consecutivos": [rg_df["Consecutivos"].sum()],
                                "Con Suplementación": [rg_df["Con Suplementación"].sum()],
                                "Con Sesion Demostrativa": [rg_df["Con Sesion Demostrativa"].sum()],
                                "Con Tamizaje 6 Meses": [rg_df["Con Tamizaje 6 Meses"].sum()],
                                "Con Anemia 6 Meses": [rg_df["Con Anemia 6 Meses"].sum()],
                                "Con Tamizaje 09 Meses": [rg_df["Con Tamizaje 09 Meses"].sum()],
                                "Con Anemia 09 Meses": [rg_df["Con Anemia 09 Meses"].sum()],
                                #"Prematuros": [rg_df["Prematuros"].sum()],
                                #"Sin Anemia (Tamizaje)": [rg_df["Sin Anemia (Tamizaje)"].sum()],
                                #"Con Anemia (Tamizaje)": [rg_df["Con Anemia (Tamizaje)"].sum()],
                            # "Sin Anemia (Diagnostico)": [rg_df["Sin Anemia (Diagnostico)"].sum()],
                            # "Con Anemia (Diagnostico)": [rg_df["Con Anemia (Diagnostico)"].sum()],
                            })
                        elif selection_rango == "120-149 días" or selection_rango == "180-209 días":
                            rg_df = rg_df.groupby(["Establecimiento de Salud"]).agg(
                                Niños_Programados=("Número de Documento", "count"),
                                Niños_Encontrados=("Estado Niño", lambda x: (x.isin(["Visita Domiciliaria (6 a 12 Meses)", "Visita Domiciliaria (1 a 5 meses)"])).sum()),
                                #Niños_Consecutivos=("¿Es consecutivo?", lambda x: (x.isin(["Consecutivo"])).sum()),
                                Con_Suplementacion=("Tipo de SUPLEMENTO", lambda x: (x.isin([
                                'MMN', 'MICRONUTRIENTES', 'GOTAS', 'FERRIMAX',
                                    'SULFATO FERROSO', 'MN', 'JARABE',
                                    'FERRAMIN FORTE', 'POLIMALTOSADO', 'SI',
                                    'FERRAMIN', '7 GOTAS', '8 GOTAS', '8 GOTITAS DE HIERRO',
                                    'MULTIMICRONUTRIENTES', 'GOTA', 'MALTOFER', 'SULFATOFERROSO',
                                    '1 MICRONUTRIENTE', 'M', 'FERRANIN FORTE',
                                    '1 SOBRECITO DE HIERRO', 'FERANIN', 'HIERRO EN GOTAS',
                                    '9 GOTITAS DE HIERRO', 'HIERRO POLIMALTOSA',
                                    'MULTIVITAMINAS', '7 GOTITAS DE HIERRO',  'FERROCIN',
                                    'GOTAS- SF','gotas', '6 GOTAS',
                                    'HIERRO DE ZINC', 'FERROCIL', 'FERROMIL'
                                ])).sum()),
                                Con_Sesion_Demostrativa=("¿Fue parte de una Sesion demostrativa?", lambda x: (x.isin(["SI"])).sum()),
                                Con_Tamizaje_6_Meses=("Resultado de Hemoglobina de 06 MESES", lambda x: (pd.to_numeric(x, errors='coerce') > 0).sum()),
                                Con_Tamizaje_6_Meses_anemia=("Resultado de Hemoglobina de 06 MESES", lambda x: ((pd.to_numeric(x, errors='coerce') < 10.5) & (pd.to_numeric(x, errors='coerce') > 0)).sum()),
                                
                                
                                #Con_Tamizaje_12_Meses=("Resultado de Hemoglobina de 12 MESES", lambda x: (pd.to_numeric(x, errors='coerce') > 0).sum()),
                                #Con_Tamizaje_12_Meses_anemia=("Resultado de Hemoglobina de 12 MESES", lambda x: ((pd.to_numeric(x, errors='coerce') < 10.5) & (pd.to_numeric(x, errors='coerce') > 0)).sum()),

                            ).reset_index()
                            rg_df.columns = ["Establecimiento de Salud","Programados", "Encontrados","Con Suplementación","Con Sesion Demostrativa","Con Tamizaje 6 Meses","Con Anemia 6 Meses"]#,"Sin Anemia (Tamizaje)","Con Anemia (Tamizaje)","Sin Anemia (Diagnostico)","Con Anemia (Diagnostico)"
                            rg_df = rg_df.sort_values("Programados", ascending=False)
                            total_row = pd.DataFrame({
                                "Establecimiento de Salud": ["TOTAL"],
                                "Programados": [rg_df["Programados"].sum()],
                                "Encontrados": [rg_df["Encontrados"].sum()],
                                #"Consecutivos": [rg_df["Consecutivos"].sum()],
                                "Con Suplementación": [rg_df["Con Suplementación"].sum()],
                                "Con Sesion Demostrativa": [rg_df["Con Sesion Demostrativa"].sum()],
                                "Con Tamizaje 6 Meses": [rg_df["Con Tamizaje 6 Meses"].sum()],
                                "Con Anemia 6 Meses": [rg_df["Con Anemia 6 Meses"].sum()],
                                #"Con Tamizaje 12 Meses": [rg_df["Con Tamizaje 12 Meses"].sum()],
                                #"Con Anemia 12 Meses": [rg_df["Con Anemia 12 Meses"].sum()],
                                #"Prematuros": [rg_df["Prematuros"].sum()],
                                #"Sin Anemia (Tamizaje)": [rg_df["Sin Anemia (Tamizaje)"].sum()],
                                #"Con Anemia (Tamizaje)": [rg_df["Con Anemia (Tamizaje)"].sum()],
                            # "Sin Anemia (Diagnostico)": [rg_df["Sin Anemia (Diagnostico)"].sum()],
                            # "Con Anemia (Diagnostico)": [rg_df["Con Anemia (Diagnostico)"].sum()],
                            })
                        else:
                            rg_df = rg_df.groupby(["Establecimiento de Salud"]).agg(
                                Niños_Programados=("Número de Documento", "count"),
                                Niños_Encontrados=("Estado Niño", lambda x: (x.isin(["Visita Domiciliaria (6 a 12 Meses)", "Visita Domiciliaria (1 a 5 meses)"])).sum()),
                                #Niños_Consecutivos=("¿Es consecutivo?", lambda x: (x.isin(["Consecutivo"])).sum()),
                                Con_Suplementacion=("Tipo de SUPLEMENTO", lambda x: (x.isin([
                                'MMN', 'MICRONUTRIENTES', 'GOTAS', 'FERRIMAX',
                                    'SULFATO FERROSO', 'MN', 'JARABE',
                                    'FERRAMIN FORTE', 'POLIMALTOSADO', 'SI',
                                    'FERRAMIN', '7 GOTAS', '8 GOTAS', '8 GOTITAS DE HIERRO',
                                    'MULTIMICRONUTRIENTES', 'GOTA', 'MALTOFER', 'SULFATOFERROSO',
                                    '1 MICRONUTRIENTE', 'M', 'FERRANIN FORTE',
                                    '1 SOBRECITO DE HIERRO', 'FERANIN', 'HIERRO EN GOTAS',
                                    '9 GOTITAS DE HIERRO', 'HIERRO POLIMALTOSA',
                                    'MULTIVITAMINAS', '7 GOTITAS DE HIERRO',  'FERROCIN',
                                    'GOTAS- SF','gotas', '6 GOTAS',
                                    'HIERRO DE ZINC', 'FERROCIL', 'FERROMIL'
                                ])).sum()),
                                Con_Sesion_Demostrativa=("¿Fue parte de una Sesion demostrativa?", lambda x: (x.isin(["SI"])).sum()),
                                Con_Tamizaje_6_Meses=("Resultado de Hemoglobina de 06 MESES", lambda x: (pd.to_numeric(x, errors='coerce') > 0).sum()),
                                Con_Tamizaje_6_Meses_anemia=("Resultado de Hemoglobina de 06 MESES", lambda x: ((pd.to_numeric(x, errors='coerce') < 10.5) & (pd.to_numeric(x, errors='coerce') > 0)).sum()),
                                
                                
                                Con_Tamizaje_12_Meses=("Resultado de Hemoglobina de 12 MESES", lambda x: (pd.to_numeric(x, errors='coerce') > 0).sum()),
                                Con_Tamizaje_12_Meses_anemia=("Resultado de Hemoglobina de 12 MESES", lambda x: ((pd.to_numeric(x, errors='coerce') < 10.5) & (pd.to_numeric(x, errors='coerce') > 0)).sum()),

                            ).reset_index()
                            rg_df.columns = ["Establecimiento de Salud","Programados", "Encontrados","Con Suplementación","Con Sesion Demostrativa","Con Tamizaje 6 Meses","Con Anemia 6 Meses","Con Tamizaje 12 Meses","Con Anemia 12 Meses"]#,"Sin Anemia (Tamizaje)","Con Anemia (Tamizaje)","Sin Anemia (Diagnostico)","Con Anemia (Diagnostico)"
                            rg_df = rg_df.sort_values("Programados", ascending=False)
                            total_row = pd.DataFrame({
                                "Establecimiento de Salud": ["TOTAL"],
                                "Programados": [rg_df["Programados"].sum()],
                                "Encontrados": [rg_df["Encontrados"].sum()],
                                #"Consecutivos": [rg_df["Consecutivos"].sum()],
                                "Con Suplementación": [rg_df["Con Suplementación"].sum()],
                                "Con Sesion Demostrativa": [rg_df["Con Sesion Demostrativa"].sum()],
                                "Con Tamizaje 6 Meses": [rg_df["Con Tamizaje 6 Meses"].sum()],
                                "Con Anemia 6 Meses": [rg_df["Con Anemia 6 Meses"].sum()],
                                "Con Tamizaje 12 Meses": [rg_df["Con Tamizaje 12 Meses"].sum()],
                                "Con Anemia 12 Meses": [rg_df["Con Anemia 12 Meses"].sum()],
                                #"Prematuros": [rg_df["Prematuros"].sum()],
                                #"Sin Anemia (Tamizaje)": [rg_df["Sin Anemia (Tamizaje)"].sum()],
                                #"Con Anemia (Tamizaje)": [rg_df["Con Anemia (Tamizaje)"].sum()],
                            # "Sin Anemia (Diagnostico)": [rg_df["Sin Anemia (Diagnostico)"].sum()],
                            # "Con Anemia (Diagnostico)": [rg_df["Con Anemia (Diagnostico)"].sum()],
                            })
                        rg_df = pd.concat([rg_df, total_row], ignore_index=True)
                        rg_df["% Encontrados"] = ((rg_df["Encontrados"] / rg_df["Programados"]) * 100).round(1).astype(str) + "%"
                        #rg_df["% SD (38%)"] = ((rg_df["Con Sesion Demostrativa"] / rg_df["Programados"]) * 100).round(1).astype(str) + "%"
                        if selection_rango == "120-149 días":
                            pass
                        elif selection_rango == "180-209 días":
                            rg_df["% Tamizajes 6 meses"] = ((rg_df["Con Tamizaje 6 Meses"] / rg_df["Programados"]) * 100).round(1).astype(str) + "%"
                        elif selection_rango == "270-299 días":
                            rg_df["% Tamizajes 09 meses"] = ((rg_df["Con Tamizaje 09 Meses"] / rg_df["Programados"]) * 100).round(1).astype(str) + "%"
                        elif selection_rango == "360-389 días":
                            rg_df["% Tamizajes 12 meses"] = ((rg_df["Con Tamizaje 12 Meses"] / rg_df["Programados"]) * 100).round(1).astype(str) + "%"
                        elif selection_rango == "6-12 meses":
                            rg_df["% Tamizajes 6 meses"] = ((rg_df["Con Tamizaje 6 Meses"] / rg_df["Programados"]) * 100).round(1).astype(str) + "%"
                            rg_df["% Tamizajes 12 meses"] = ((rg_df["Con Tamizaje 12 Meses"] / rg_df["Programados"]) * 100).round(1).astype(str) + "%"
                        #rg_df["% Tamizaje"] = ((rg_df["Sin Anemia (Tamizaje)"] / rg_df["Programados"]) * 100).round(1).astype(str) + "%"
                        #rg_df["% Diagnostico"] = ((rg_df["Sin Anemia (Diagnostico)"] / rg_df["Programados"]) * 100).round(1).astype(str) + "%"
                        gb = GridOptionsBuilder.from_dataframe(rg_df)
                        gb.configure_default_column(cellStyle={'fontSize': '17px'}) 
                        gb.configure_column("Establecimiento de Salud", width=400, pinned='left')
                        
                        # Configurar ancho de columnas numéricas
                        #numeric_columns = ["Programados", "Encontrados","Con Suplementación","Sin Anemia (Tamizaje)","Con Anemia (Tamizaje)","Sin Anemia (Diagnostico)","Con Anemia (Diagnostico)",]
                        #for col in numeric_columns:
                        #     gb.configure_column(col, width=130)
                        
                        grid_options = gb.build()
                        #grid_options['getRowStyle'] = row_style
                        
                        grid_response = AgGrid(rg_df, # Dataframe a mostrar
                                                gridOptions=grid_options,
                                                enable_enterprise_modules=False,
                                                #theme='balham',  # Cambiar tema si se desea ('streamlit', 'light', 'dark', 'alpine', etc.)
                                                update_mode='MODEL_CHANGED',
                                                #fit_columns_on_grid_load=True,
                                                height=370,
                                                key="grid_anemia"
                                            )
                        
                        # Función para construir dataframe para cada rango de edad
                        def build_rg_df_for(selection_label):
                            if selection_label == "120-149 días":
                                temp_df = dataframe_[dataframe_["Rango de Días Activo"]=="Rango de días 120-149 días"]
                            elif selection_label == "180-209 días":
                                temp_df = dataframe_[dataframe_["Rango de Días Activo"]=="Rango de días 180-209 días"]
                            elif selection_label == "270-299 días":
                                temp_df = dataframe_[dataframe_["Rango de Días Activo"]=="Rango de días 270-299 días"]
                                # Asegura columna de Hb 09 meses si falta para evitar KeyError
                                if "Resultado de Hemoglobina de 09 MESES" not in temp_df.columns:
                                    temp_df = temp_df.copy()
                                    temp_df["Resultado de Hemoglobina de 09 MESES"] = 0
                            elif selection_label == "360-389 días":
                                temp_df = dataframe_[dataframe_["Rango de Días Activo"]=="Rango de días 360-389 días"]
                            elif selection_label == "6-12 meses":
                                temp_df = dataframe_[dataframe_["Rango de Edad"]=="6-12 meses"]
                            
                            if selection_label == "270-299 días":
                                temp_df_grouped = temp_df.groupby(["Establecimiento de Salud"]).agg(
                                    Niños_Programados=("Número de Documento", "count"),
                                    Niños_Encontrados=("Estado Niño", lambda x: (x.isin(["Visita Domiciliaria (6 a 12 Meses)", "Visita Domiciliaria (1 a 5 meses)"])).sum()),
                                    Con_Suplementacion=("Tipo de SUPLEMENTO", lambda x: (x.isin([
                                    'MMN', 'MICRONUTRIENTES', 'GOTAS', 'FERRIMAX',
                                        'SULFATO FERROSO', 'MN', 'JARABE',
                                        'FERRAMIN FORTE', 'POLIMALTOSADO', 'SI',
                                        'FERRAMIN', '7 GOTAS', '8 GOTAS', '8 GOTITAS DE HIERRO',
                                        'MULTIMICRONUTRIENTES', 'GOTA', 'MALTOFER', 'SULFATOFERROSO',
                                        '1 MICRONUTRIENTE', 'M', 'FERRANIN FORTE',
                                        '1 SOBRECITO DE HIERRO', 'FERANIN', 'HIERRO EN GOTAS',
                                        '9 GOTITAS DE HIERRO', 'HIERRO POLIMALTOSA',
                                        'MULTIVITAMINAS', '7 GOTITAS DE HIERRO',  'FERROCIN',
                                        'GOTAS- SF','gotas', '6 GOTAS',
                                        'HIERRO DE ZINC', 'FERROCIL', 'FERROMIL'
                                    ])).sum()),
                                    Con_Sesion_Demostrativa=("¿Fue parte de una Sesion demostrativa?", lambda x: (x.isin(["SI"])).sum()),
                                    Con_Tamizaje_6_Meses=("Resultado de Hemoglobina de 06 MESES", lambda x: (pd.to_numeric(x, errors='coerce') > 0).sum()),
                                    Con_Tamizaje_6_Meses_anemia=("Resultado de Hemoglobina de 06 MESES", lambda x: ((pd.to_numeric(x, errors='coerce') < 10.5) & (pd.to_numeric(x, errors='coerce') > 0)).sum()),
                                    Con_Tamizaje_09_Meses=("Resultado de Hemoglobina de 09 MESES", lambda x: (pd.to_numeric(x, errors='coerce') > 0).sum()),
                                    Con_Tamizaje_09_Meses_anemia=("Resultado de Hemoglobina de 09 MESES", lambda x: ((pd.to_numeric(x, errors='coerce') < 10.5) & (pd.to_numeric(x, errors='coerce') > 0)).sum()),
                                ).reset_index()
                                temp_df_grouped.columns = ["Establecimiento de Salud","Programados", "Encontrados","Con Suplementación","Con Sesion Demostrativa","Con Tamizaje 6 Meses","Con Anemia 6 Meses","Con Tamizaje 09 Meses","Con Anemia 09 Meses"]
                                temp_df_grouped = temp_df_grouped.sort_values("Programados", ascending=False)
                                total_row = pd.DataFrame({
                                    "Establecimiento de Salud": ["TOTAL"],
                                    "Programados": [temp_df_grouped["Programados"].sum()],
                                    "Encontrados": [temp_df_grouped["Encontrados"].sum()],
                                    "Con Suplementación": [temp_df_grouped["Con Suplementación"].sum()],
                                    "Con Sesion Demostrativa": [temp_df_grouped["Con Sesion Demostrativa"].sum()],
                                    "Con Tamizaje 6 Meses": [temp_df_grouped["Con Tamizaje 6 Meses"].sum()],
                                    "Con Anemia 6 Meses": [temp_df_grouped["Con Anemia 6 Meses"].sum()],
                                    "Con Tamizaje 09 Meses": [temp_df_grouped["Con Tamizaje 09 Meses"].sum()],
                                    "Con Anemia 09 Meses": [temp_df_grouped["Con Anemia 09 Meses"].sum()],
                                })
                                temp_df_grouped = pd.concat([temp_df_grouped, total_row], ignore_index=True)
                                temp_df_grouped["% Encontrados"] = ((temp_df_grouped["Encontrados"] / temp_df_grouped["Programados"]) * 100).round(1).astype(str) + "%"
                                temp_df_grouped["% Tamizajes 09 meses"] = ((temp_df_grouped["Con Tamizaje 09 Meses"] / temp_df_grouped["Programados"]) * 100).round(1).astype(str) + "%"
                                
                            elif selection_label == "120-149 días" or selection_label == "180-209 días":
                                temp_df_grouped = temp_df.groupby(["Establecimiento de Salud"]).agg(
                                    Niños_Programados=("Número de Documento", "count"),
                                    Niños_Encontrados=("Estado Niño", lambda x: (x.isin(["Visita Domiciliaria (6 a 12 Meses)", "Visita Domiciliaria (1 a 5 meses)"])).sum()),
                                    Con_Suplementacion=("Tipo de SUPLEMENTO", lambda x: (x.isin([
                                    'MMN', 'MICRONUTRIENTES', 'GOTAS', 'FERRIMAX',
                                        'SULFATO FERROSO', 'MN', 'JARABE',
                                        'FERRAMIN FORTE', 'POLIMALTOSADO', 'SI',
                                        'FERRAMIN', '7 GOTAS', '8 GOTAS', '8 GOTITAS DE HIERRO',
                                        'MULTIMICRONUTRIENTES', 'GOTA', 'MALTOFER', 'SULFATOFERROSO',
                                        '1 MICRONUTRIENTE', 'M', 'FERRANIN FORTE',
                                        '1 SOBRECITO DE HIERRO', 'FERANIN', 'HIERRO EN GOTAS',
                                        '9 GOTITAS DE HIERRO', 'HIERRO POLIMALTOSA',
                                        'MULTIVITAMINAS', '7 GOTITAS DE HIERRO',  'FERROCIN',
                                        'GOTAS- SF','gotas', '6 GOTAS',
                                        'HIERRO DE ZINC', 'FERROCIL', 'FERROMIL'
                                    ])).sum()),
                                    Con_Sesion_Demostrativa=("¿Fue parte de una Sesion demostrativa?", lambda x: (x.isin(["SI"])).sum()),
                                    Con_Tamizaje_6_Meses=("Resultado de Hemoglobina de 06 MESES", lambda x: (pd.to_numeric(x, errors='coerce') > 0).sum()),
                                    Con_Tamizaje_6_Meses_anemia=("Resultado de Hemoglobina de 06 MESES", lambda x: ((pd.to_numeric(x, errors='coerce') < 10.5) & (pd.to_numeric(x, errors='coerce') > 0)).sum()),
                                ).reset_index()
                                temp_df_grouped.columns = ["Establecimiento de Salud","Programados", "Encontrados","Con Suplementación","Con Sesion Demostrativa","Con Tamizaje 6 Meses","Con Anemia 6 Meses"]
                                temp_df_grouped = temp_df_grouped.sort_values("Programados", ascending=False)
                                total_row = pd.DataFrame({
                                    "Establecimiento de Salud": ["TOTAL"],
                                    "Programados": [temp_df_grouped["Programados"].sum()],
                                    "Encontrados": [temp_df_grouped["Encontrados"].sum()],
                                    "Con Suplementación": [temp_df_grouped["Con Suplementación"].sum()],
                                    "Con Sesion Demostrativa": [temp_df_grouped["Con Sesion Demostrativa"].sum()],
                                    "Con Tamizaje 6 Meses": [temp_df_grouped["Con Tamizaje 6 Meses"].sum()],
                                    "Con Anemia 6 Meses": [temp_df_grouped["Con Anemia 6 Meses"].sum()],
                                })
                                temp_df_grouped = pd.concat([temp_df_grouped, total_row], ignore_index=True)
                                temp_df_grouped["% Encontrados"] = ((temp_df_grouped["Encontrados"] / temp_df_grouped["Programados"]) * 100).round(1).astype(str) + "%"
                                if selection_label == "180-209 días":
                                    temp_df_grouped["% Tamizajes 6 meses"] = ((temp_df_grouped["Con Tamizaje 6 Meses"] / temp_df_grouped["Programados"]) * 100).round(1).astype(str) + "%"
                                    
                            else:  # 360-389 días o 6-12 meses
                                temp_df_grouped = temp_df.groupby(["Establecimiento de Salud"]).agg(
                                    Niños_Programados=("Número de Documento", "count"),
                                    Niños_Encontrados=("Estado Niño", lambda x: (x.isin(["Visita Domiciliaria (6 a 12 Meses)", "Visita Domiciliaria (1 a 5 meses)"])).sum()),
                                    Con_Suplementacion=("Tipo de SUPLEMENTO", lambda x: (x.isin([
                                    'MMN', 'MICRONUTRIENTES', 'GOTAS', 'FERRIMAX',
                                        'SULFATO FERROSO', 'MN', 'JARABE',
                                        'FERRAMIN FORTE', 'POLIMALTOSADO', 'SI',
                                        'FERRAMIN', '7 GOTAS', '8 GOTAS', '8 GOTITAS DE HIERRO',
                                        'MULTIMICRONUTRIENTES', 'GOTA', 'MALTOFER', 'SULFATOFERROSO',
                                        '1 MICRONUTRIENTE', 'M', 'FERRANIN FORTE',
                                        '1 SOBRECITO DE HIERRO', 'FERANIN', 'HIERRO EN GOTAS',
                                        '9 GOTITAS DE HIERRO', 'HIERRO POLIMALTOSA',
                                        'MULTIVITAMINAS', '7 GOTITAS DE HIERRO',  'FERROCIN',
                                        'GOTAS- SF','gotas', '6 GOTAS',
                                        'HIERRO DE ZINC', 'FERROCIL', 'FERROMIL'
                                    ])).sum()),
                                    Con_Sesion_Demostrativa=("¿Fue parte de una Sesion demostrativa?", lambda x: (x.isin(["SI"])).sum()),
                                    Con_Tamizaje_6_Meses=("Resultado de Hemoglobina de 06 MESES", lambda x: (pd.to_numeric(x, errors='coerce') > 0).sum()),
                                    Con_Tamizaje_6_Meses_anemia=("Resultado de Hemoglobina de 06 MESES", lambda x: ((pd.to_numeric(x, errors='coerce') < 10.5) & (pd.to_numeric(x, errors='coerce') > 0)).sum()),
                                    Con_Tamizaje_12_Meses=("Resultado de Hemoglobina de 12 MESES", lambda x: (pd.to_numeric(x, errors='coerce') > 0).sum()),
                                    Con_Tamizaje_12_Meses_anemia=("Resultado de Hemoglobina de 12 MESES", lambda x: ((pd.to_numeric(x, errors='coerce') < 10.5) & (pd.to_numeric(x, errors='coerce') > 0)).sum()),
                                ).reset_index()
                                temp_df_grouped.columns = ["Establecimiento de Salud","Programados", "Encontrados","Con Suplementación","Con Sesion Demostrativa","Con Tamizaje 6 Meses","Con Anemia 6 Meses","Con Tamizaje 12 Meses","Con Anemia 12 Meses"]
                                temp_df_grouped = temp_df_grouped.sort_values("Programados", ascending=False)
                                total_row = pd.DataFrame({
                                    "Establecimiento de Salud": ["TOTAL"],
                                    "Programados": [temp_df_grouped["Programados"].sum()],
                                    "Encontrados": [temp_df_grouped["Encontrados"].sum()],
                                    "Con Suplementación": [temp_df_grouped["Con Suplementación"].sum()],
                                    "Con Sesion Demostrativa": [temp_df_grouped["Con Sesion Demostrativa"].sum()],
                                    "Con Tamizaje 6 Meses": [temp_df_grouped["Con Tamizaje 6 Meses"].sum()],
                                    "Con Anemia 6 Meses": [temp_df_grouped["Con Anemia 6 Meses"].sum()],
                                    "Con Tamizaje 12 Meses": [temp_df_grouped["Con Tamizaje 12 Meses"].sum()],
                                    "Con Anemia 12 Meses": [temp_df_grouped["Con Anemia 12 Meses"].sum()],
                                })
                                temp_df_grouped = pd.concat([temp_df_grouped, total_row], ignore_index=True)
                                temp_df_grouped["% Encontrados"] = ((temp_df_grouped["Encontrados"] / temp_df_grouped["Programados"]) * 100).round(1).astype(str) + "%"
                                if selection_label == "360-389 días":
                                    temp_df_grouped["% Tamizajes 12 meses"] = ((temp_df_grouped["Con Tamizaje 12 Meses"] / temp_df_grouped["Programados"]) * 100).round(1).astype(str) + "%"
                                elif selection_label == "6-12 meses":
                                    temp_df_grouped["% Tamizajes 6 meses"] = ((temp_df_grouped["Con Tamizaje 6 Meses"] / temp_df_grouped["Programados"]) * 100).round(1).astype(str) + "%"
                                    temp_df_grouped["% Tamizajes 12 meses"] = ((temp_df_grouped["Con Tamizaje 12 Meses"] / temp_df_grouped["Programados"]) * 100).round(1).astype(str) + "%"
                            
                            return temp_df_grouped
                        
                        # Generar Excel con todas las hojas y formato profesional
                        def generate_excel_all_ranges():
                            from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
                            from openpyxl.utils.dataframe import dataframe_to_rows
                            
                            output = BytesIO()
                            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                                age_ranges = ["120-149 días", "180-209 días", "270-299 días", "360-389 días", "6-12 meses"]
                                
                                for age_range in age_ranges:
                                    df_range = build_rg_df_for(age_range)
                                    # Limpiar nombre de hoja para Excel (remover caracteres especiales)
                                    sheet_name = age_range.replace("-", "_").replace(" ", "_")
                                    df_range.to_excel(writer, sheet_name=sheet_name, index=False)
                                    
                                    # Obtener la hoja de trabajo para aplicar formato
                                    worksheet = writer.sheets[sheet_name]
                                    
                                    # Definir estilos
                                    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                                    header_font = Font(color="FFFFFF", bold=True, size=11)
                                    
                                    total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                                    total_font = Font(bold=True, size=11)
                                    
                                    border = Border(
                                        left=Side(style='thin'),
                                        right=Side(style='thin'),
                                        top=Side(style='thin'),
                                        bottom=Side(style='thin')
                                    )
                                    
                                    center_alignment = Alignment(horizontal='center', vertical='center')
                                    left_alignment = Alignment(horizontal='left', vertical='center')
                                    
                                    # Aplicar formato a los headers (fila 1)
                                    for col_num in range(1, len(df_range.columns) + 1):
                                        cell = worksheet.cell(row=1, column=col_num)
                                        cell.fill = header_fill
                                        cell.font = header_font
                                        cell.alignment = center_alignment
                                        cell.border = border
                                    
                                    # Aplicar formato a las celdas de datos
                                    for row_num in range(2, len(df_range) + 2):
                                        for col_num in range(1, len(df_range.columns) + 1):
                                            cell = worksheet.cell(row=row_num, column=col_num)
                                            cell.border = border
                                            
                                            # Alineación: centrada para números, izquierda para texto
                                            if col_num == 1:  # Columna de Establecimiento
                                                cell.alignment = left_alignment
                                            else:
                                                cell.alignment = center_alignment
                                            
                                            # Formato especial para la fila TOTAL
                                            if df_range.iloc[row_num-2, 0] == "TOTAL":
                                                cell.fill = total_fill
                                                cell.font = total_font
                                    
                                    # Ajustar ancho de columnas
                                    for col_num, column in enumerate(df_range.columns, 1):
                                        column_letter = get_column_letter(col_num)
                                        if col_num == 1:  # Columna de Establecimiento
                                            worksheet.column_dimensions[column_letter].width = 35
                                        else:
                                            worksheet.column_dimensions[column_letter].width = 15
                                    
                                    # Congelar la primera fila
                                    worksheet.freeze_panes = 'A2'
                            
                            output.seek(0)
                            return output.getvalue()
                        
                        # Botón de descarga
                        excel_data = generate_excel_all_ranges()
                        st.download_button(
                            label="📥 Descargar Excel (todos los rangos)",
                            data=excel_data,
                            file_name=f"seguimiento_indicadores_SN_{select_mes}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                    
                        #tamizaje_df = rg_dff.groupby(["ESTADO TAMIZAJE"])[["Tipo Documento(P)"]].count().reset_index()
                        #tamizaje_df.columns = ["ESTADO TAMIZAJE", "Niños"]
                        #diagnostico_df = rg_dff.groupby(["ESTADO HB NIÑO"])[["Tipo Documento(P)"]].count().reset_index()
                        #diagnostico_df.columns = ["ESTADO HB NIÑO", "Niños"]

                        #col1, col2 = st.columns(2)
                        #with col1:
                        #    fig_tamizaje = px.pie(tamizaje_df, values='Niños', names='ESTADO TAMIZAJE', title='Distribución de Resultados de Tamizaje')
                        #    fig_tamizaje.update_traces(textposition='inside', textinfo='percent+label+value', insidetextfont=dict(size=18))
                        #    st.plotly_chart(fig_tamizaje)
                        #with col2:
                        #    fig_diagnostico = px.pie(diagnostico_df, values='Niños', names='ESTADO HB NIÑO', title='Distribución de Resultados de Diagnóstico')
                        #    fig_diagnostico.update_traces(textposition='inside', textinfo='percent+label+value', insidetextfont=dict(size=18))
                        #    st.plotly_chart(fig_diagnostico)
        

                    render_anemia_indicators()
                """
                # Filtra duplicados de Celular Madre
                df_con_num_cel = dataframe_[dataframe_["Celular Madre"] != 0].copy()
                df_con_num_cel["Celular Madre"] = df_con_num_cel["Celular Madre"].astype(str)

                    # Detecta duplicados de Número Doc Madre
                duplicados_num_doc_madre = df_con_num_cel["Número Doc Madre"].duplicated(keep=False)

                    # Filtra duplicados de Celular Madre y excluye duplicados de Número Doc Madre
                cel_duplicados_df = df_con_num_cel[
                        df_con_num_cel.duplicated(subset=["Celular Madre"], keep=False) &
                        ~duplicados_num_doc_madre
                    ]
                    

                groupings = [
                        # (df, x, y, title, text, orientation, color, barmode, xaxis_title, yaxis_title)
                        (dataframe_.groupby(['Establecimiento de Salud'])[['Número de Documento']].count()
                            .rename(columns={"Número de Documento":"Registros"})
                            .sort_values("Registros", ascending=True).reset_index(),
                        "Registros", "Establecimiento de Salud", "Niños Asignados por Establecimiento de Salud", "Registros", 'h', None, None, "Número de Niños Cargados", None),
                        (dataframe_.groupby(['Establecimiento de Salud'])[['Total de VD presenciales Válidas']].sum()
                            .rename(columns={"Total de VD presenciales Válidas":"Visitas"})
                            .sort_values("Visitas", ascending=True).reset_index(),
                        "Visitas", "Establecimiento de Salud", "Niños Visitados por Establecimiento de Salud", "Visitas", 'h', None, None, "Número de Visitas", None),
                        (dataframe_.groupby(["Estado Niño"])[["Tipo Documento"]].count().rename(columns={"Tipo Documento":"Niños"}).reset_index(),
                        "Niños", "Estado Niño", f"Estado de los Niños Cargados {select_mes}", "Niños", 'h', None, None, "Niños", None),
                        (dataframe_.groupby(["Entidad Actualiza"])[["Tipo Documento"]].count().rename(columns={"Tipo Documento":"Niños"}).reset_index(),
                        "Niños", "Entidad Actualiza", f"Niños C1 - Estado Actualizaciones Padrón Nominal {select_mes}", "Niños", 'h', None, None, "Niños", None),
                        (dataframe_.groupby(["Estado Visitas"])[["Tipo Documento"]].count().rename(columns={"Tipo Documento":"Niños"}).reset_index(),
                        "Niños", "Estado Visitas", f"Estado de las Visitas en {select_mes}", "Niños", 'h', None, None, "Niños", None),
                        (dataframe_.groupby(['Establecimiento de Salud'])[['Total de VD presencial Válidas MOVIL']].sum()
                            .rename(columns={"Total de VD presencial Válidas MOVIL":"Visitas"})
                            .sort_values("Visitas", ascending=True).reset_index(),
                        "Visitas", "Establecimiento de Salud", "Visitas Registradas por MOVIL", "Visitas", 'h', None, None, "Número de Visitas", None),
                        (dataframe_.groupby(['Establecimiento de Salud'])[['Total de VD presencial Válidas WEB']].sum()
                            .rename(columns={"Total de VD presencial Válidas WEB":"Visitas"})
                            .sort_values("Visitas", ascending=True).reset_index(),
                        "Visitas", "Establecimiento de Salud", "Visitas Registradas por WEB", "Visitas", 'h', None, None, "Número de Visitas", None),
                    ]

                    # Gráficos de barras
                figs = [plot_bar(*args) for args in groupings]

                    # Gráficos especiales
                    # Niños No Encontrados
                no_encontrados_df = dataframe_[dataframe_["Estado Niño"]=="No Encontrado"] \
                        .groupby(["Establecimiento de Salud"])[["Estado Niño"]].count() \
                        .rename(columns={"Estado Niño":"Registros"}).sort_values("Registros", ascending=True).reset_index()
                fig_eess_count_noencon = plot_bar(no_encontrados_df, "Registros", "Establecimiento de Salud", "Niños No Encontrados por Establecimiento de Salud", "Registros", 'h', None, None, "Número de Niños No Encontrados", None)

                    # Niños Rechazados
                rechazado_df = dataframe_[dataframe_["Estado Niño"]=="Rechazado"] \
                        .groupby(["Establecimiento de Salud"])[["Estado Niño"]].count() \
                        .rename(columns={"Estado Niño":"Registros"}).sort_values("Registros", ascending=True).reset_index()
                fig_eess_count_rechazado = plot_bar(rechazado_df, "Registros", "Establecimiento de Salud", "Niños Rechazados por Establecimiento de Salud", "Registros", 'h', None, None, "Número de Niños Rechazados", None)

                    # Niños sin Visita
                sinvd_dff = dataframe_[dataframe_["Estado Niño"]==f"Sin Visita ({select_mes})"]
                sinvd_df = sinvd_dff.groupby(["Establecimiento de Salud"])[["Estado Niño"]].count() \
                        .rename(columns={"Estado Niño":"Niños"}).sort_values("Niños", ascending=True).reset_index()
                fig_eess_sinvd = plot_bar(sinvd_df, "Niños", "Establecimiento de Salud", "Niños sin Visita por Establecimiento", "Niños", 'h', None, None, "Número de Niños", None)

                    # Niños Transito
                child_transito_df = dataframe_[dataframe_["Tipo Registro Padrón Nominal"]=="Activos Transito"] \
                        .groupby(["Establecimiento de Salud"])[["Tipo Registro Padrón Nominal"]].count() \
                        .rename(columns={"Tipo Registro Padrón Nominal":"Niños"}).sort_values("Niños", ascending=True).reset_index()
                fig_eess_transito = plot_bar(child_transito_df, "Niños", "Establecimiento de Salud", "Niños Transito por Establecimiento", "Niños", 'h', None, None, "Número de Niños", None)

                    # Niños Otro Padrón
                child_deri_df = dataframe_[dataframe_["Tipo Registro Padrón Nominal"]=="En Otro Padrón Nominal"] \
                        .groupby(["Establecimiento de Salud"])[["Tipo Registro Padrón Nominal"]].count() \
                        .rename(columns={"Tipo Registro Padrón Nominal":"Niños"}).sort_values("Niños", ascending=True).reset_index()
                fig_eess_derivados = plot_bar(child_deri_df, "Niños", "Establecimiento de Salud", "Niños En otro Distrito Padrón por Establecimiento", "Niños", 'h', None, None, "Número de Niños", None)

                    # Estado de Derivados y Transitos
                registro_padron_df = dataframe_[dataframe_["Tipo Registro Padrón Nominal"].isin(["Activos Transito","En Otro Padrón Nominal"])]
                registro_padron_group_df = registro_padron_df.groupby(["Estado Niño","Tipo Registro Padrón Nominal"])[["Número de Documento"]].count().rename(columns={"Número de Documento":"Niños"}).reset_index()
                fig_reg_padron = px.bar(registro_padron_group_df, x="Estado Niño", y="Niños", color="Tipo Registro Padrón Nominal",
                                            text="Niños", orientation='v', title="Niños Derivados y Transitos por Estado de Visita", barmode='group')
                fig_reg_padron.update_traces(textfont_size=18, textangle=0, textposition="outside", cliponaxis=False)
                fig_reg_padron.update_layout(xaxis=dict(title=dict(text="Estado de Visita de los Niños")), yaxis=dict(title=dict(text="Número de Niños")), font=dict(size=16))

                    # Tipo de Seguro por Establecimiento
                tipo_seguro_df = dataframe_.groupby(["Establecimiento de Salud","Tipo de Seguro"])[["Número de Documento"]].count().rename(columns={"Número de Documento":"Niños"}).reset_index()
                fig_tipo_seguro = plot_bar(tipo_seguro_df, "Niños", "Establecimiento de Salud", "Tipo de Seguro por Establicimiento de salud", "Niños", 'h', "Tipo de Seguro", 'stack', "Número de Niños", "Establecimiento de Salud")

                    # Tipo de Seguro general
                tipo_seguro_all_df = dataframe_.groupby(["Tipo de Seguro"])[["Número de Documento"]].count().rename(columns={"Número de Documento":"Niños"}).reset_index()
                fig_seguro_all = px.pie(tipo_seguro_all_df, values='Niños', names='Tipo de Seguro', title='Tipo de Seguro')
                fig_seguro_all.update_traces(textposition='inside', textinfo='percent+label+value', insidetextfont=dict(size=18))
                    
                with st.expander("Charts"):
                        col1, col2 = st.columns(2)
                        with col1:
                            tab1, tab2, tab3, tab4,tab5 = st.tabs(["Asignados", "Visitados","Sin Visitas","Visitas Movil","Visitas Web"])
                            with tab1:
                                st.plotly_chart(figs[0], use_container_width=True)  # Niños Asignados
                            with tab2:
                                st.plotly_chart(figs[1], use_container_width=True) 
                            with tab3:
                                st.plotly_chart(fig_eess_sinvd, use_container_width=True)
                            with tab4:
                                st.plotly_chart(figs[5], use_container_width=True)
                            with tab5:
                                st.plotly_chart(figs[6], use_container_width=True)


                        with col2:
                            tab1, tab2, tab3, tab4,tab5 = st.tabs(["Transitos/Derivados","Niños Otro Padrón","Niños Transito","Niños No Encontrados", "Niños Rechazados"])
                            with tab1:
                                st.plotly_chart(fig_reg_padron, use_container_width=True)
                            with tab2:
                                st.plotly_chart(fig_eess_derivados, use_container_width=True)
                            with tab3:
                                st.plotly_chart(fig_eess_transito, use_container_width=True)
                                
                            with tab4:
                                st.plotly_chart(fig_eess_count_rechazado, use_container_width=True)  # Niños Rechazados
                            with tab5:
                                st.plotly_chart(fig_eess_count_noencon, use_container_width=True)  # Niños No Encontrados
                        
                        col1_, col2_= st.columns(2)
                        with col1_:
                            tab1_, tab2_, tab3_ = st.tabs(["Tipo de Etapa","Etapa por Visitas","Entidad Actualiza"])
                            with tab1_:
                                st.plotly_chart(figs[2], use_container_width=True)
                            with tab2_:
                                st.plotly_chart(figs[4], use_container_width=True)
                            with tab3_:
                                st.plotly_chart(figs[3], use_container_width=True)
                        with col2_: 
                            st.plotly_chart(fig_seguro_all, use_container_width=True)
                #fecha_actual
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.download_button(
                            label="Descargar Reporte Completo",
                            icon=":material/download:",
                            data=convert_excel_df(dataframe_),
                            file_name=f"EstadoVisitas_{select_year}_{select_mes}_{fecha_actual}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                with col2:
                    st.download_button( 
                            label="Descargar Reporte C1",
                            icon=":material/download:",
                            data=convert_excel_df(dataframe_[COLS_DATAFRAME_RESUMIDO_CHILS]),
                            file_name=f"EstadoVisitas_{select_year}_{select_mes}_{fecha_actual}_r.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                with col3:
                    st.download_button(
                            label="Descargar Reporte Padron Nominal",
                            icon=":material/download:",
                            data=convert_excel_df(dataframe_[COLS_DATAFRAME_PADRON_CHILS]),
                            file_name=f"EstadoVisitas_{select_year}_{select_mes}_{fecha_actual}_r.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                with col4:
                    st.download_button(
                            label="Descargar Reporte Duplicados",
                            icon=":material/download:",
                            data=convert_excel_df(cel_duplicados_df[COLS_DATAFRAME_DUPLICADOS_CHILDS]),
                            file_name=f"Duplicados_{select_year}_{select_mes}_{fecha_actual}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
            #except:
            #    st.warning("Datos incompletos para este periodo")

            
            
            



def estadisticas_dashboard():
    styles(2)
    c1_carga_df = fetch_carga_childs()
    #c1_carga_df["Número de Documento del niño"] = c1_carga_df["Número de Documento del niño"].astype(str).str.strip()
    
    supledf = pd.read_parquet(r'./data/microred/suplementacion.parquet', engine='pyarrow')
    supledf["DNI_PACIENTE"] = supledf["DNI_PACIENTE"].str.strip()
    supledf = supledf.rename(columns={"DNI_PACIENTE":"Documento"})
    supledf = supledf[(supledf["ACTIVIDAD"].str.contains("85018"))|(supledf["ACTIVIDAD"].str.contains("D5"))]
    st.write(supledf)
    
    st.title("Indicador: Niños sin anemia")
    #jun_seg_nominal_df = c1_carga_df[(c1_carga_df["Mes"] == 6) & (c1_carga_df["Año"] == 2025)]
    #jun_seg_nominal_df = jun_seg_nominal_df[[
    #    "Establecimiento de Salud","Nombres del Actor Social","Tipo de Documento del niño","Número de Documento del niño","Fecha de Nacimiento","Rango de Edad","DNI de la madre","Celular de la madre",
    #    "Mes","Año"
    #]]
    all_c1_carga_df = c1_carga_df[(c1_carga_df["Año"] == 2025)&(c1_carga_df["Mes"].isin([1,2,3,4,5,6,7]))]
    
    all_c1_carga_df["Mes_name"] = all_c1_carga_df["Mes"].apply(mes_short)
    all_c1_carga_df["Mes_name"] = all_c1_carga_df["Mes_name"] + "-"
    all_c1_carga_df = all_c1_carga_df.sort_values(by="Mes", ascending=True)
    all_c1_carga_df["Mes"] = all_c1_carga_df["Mes"].astype(str)
    #all_c1_carga_df['Edad_Meses'] = all_c1_carga_df['Fecha de Nacimiento'].apply(lambda x: (datetime.now() - x).days / 30.44)
    #all_c1_carga_df['Tiene_6_Meses'] = all_c1_carga_df['Edad_Meses'].apply(lambda x: 'Sí' if x >= 6 else 'No')
    #all_c1_carga_df.to_excel("owo.xlsx")
    
    total = all_c1_carga_df.groupby(["Número de Documento del niño"]).agg({"Mes": "sum","Mes_name": "sum"}).reset_index()
    all_c1_carga_df = all_c1_carga_df[all_c1_carga_df["Rango de Edad"].isin(["6-12 meses"])]
    
    #all_c1_carga_df = all_c1_carga_df[all_c1_carga_df["Tiene_6_Meses"]=="Sí"]
    all_c1_carga_df["Establecimiento de Salud"] = all_c1_carga_df["Establecimiento de Salud"]+" - "
    
    #st.write(len(list(all_c1_carga_df["Número de Documento del niño"].unique())))

    unique_childs25_df = all_c1_carga_df.groupby(["Número de Documento del niño"]).agg({"Mes": "sum","Mes_name": "sum","Establecimiento de Salud": "sum"}).reset_index()
    unique_childs25_df.columns = ["Documento", "Periodos","Periodos_name","Establecimiento de Salud"]
    unique_childs25_df["Periodos_name"] = unique_childs25_df["Periodos_name"].str[:-1] 
    unique_childs25_df['Es_Consecutivo'] = unique_childs25_df['Periodos'].apply(es_consecutivo)  
    unique_childs25_df = unique_childs25_df[["Documento","Periodos_name","Es_Consecutivo","Establecimiento de Salud"]]
    unique_childs25_df.columns = ["Documento","Periodos","¿Es consecutivo?","Establecimiento de Salud"]
    unique_childs25_df["Documento"] = unique_childs25_df["Documento"].astype(str)
    unique_childs25_df["Documento"] = unique_childs25_df["Documento"].str.strip()
    print(unique_childs25_df.shape)
    childs_total =total.shape[0]
    childs_consecutivos = (unique_childs25_df["¿Es consecutivo?"]=="Consecutivo").sum()
   
    #childs_no_consecutivos = (unique_childs25_df["¿Es consecutivo?"]=="No Consecutivo").sum()
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Total de Niños Cargados 2025", childs_total)
    with col2:
        st.metric("Total de Niños Consecutivos ", childs_consecutivos)#de 6 meses
    with col3:
        st.metric("Umbral Minimo 30%", round(childs_consecutivos*0.3,0))
    with col4:
        st.metric("Umbral Minimo 37%", round(childs_consecutivos*0.37,0))

    #st.write(unique_childs25_df.shape)
    #st.dataframe(unique_childs25_df)
    #st.write(supledf.shape)
    #st.dataframe(supledf)
    unique_childs25_df = unique_childs25_df[unique_childs25_df["¿Es consecutivo?"]=="Consecutivo"]
    st.write(unique_childs25_df.shape)
    st.dataframe(unique_childs25_df)
    
    # Obtener el último diagnóstico por documento
    supledf_sorted = supledf.sort_values(['Documento', 'Fecha_Diagnostico'], ascending=[True, False])
    ultimo_diagnostico_df = supledf_sorted.groupby('Documento').first().reset_index()
    ultimo_diagnostico_df = ultimo_diagnostico_df[['Documento', 'DIAGNOSTICO', 'Fecha_Diagnostico']]
    ultimo_diagnostico_df = ultimo_diagnostico_df.rename(columns={'DIAGNOSTICO': 'Ultimo Diagnostico', 'Fecha_Diagnostico': 'Fecha Ultimo Diagnostico'})
    
    # Hacer merge con unique_childs25_df para obtener el último diagnóstico
    unique_childs25_df = pd.merge(unique_childs25_df, ultimo_diagnostico_df, on="Documento", how="left")
    unique_childs25_df['Ultimo Diagnostico'] = unique_childs25_df['Ultimo Diagnostico'].fillna('SIN DIAGNOSTICO')
    
    st.write("Datos con último diagnóstico:")
    st.write(unique_childs25_df.shape)
    st.dataframe(unique_childs25_df)
    
    # Hacer merge completo para el análisis detallado
    df_ = pd.merge(unique_childs25_df, supledf, on="Documento", how="left")
    df_ = df_.sort_values(by="Fecha_Diagnostico", ascending=True)
    st.write(df_.shape)
    st.dataframe(df_)













    
    df_["MICRORED"] = df_["MICRORED"].replace("TRUJILLO - METROPOLITANO", "TRUJILLO")
    df_["PACIENTE"] = df_["PACIENTE"].fillna("Sin Datos")
    df_["PACIENTE"] = df_["PACIENTE"].str.strip()
    #print(df_["PACIENTE"].unique())
    # Calcular edad de diagnóstico si tienes fecha de nacimiento y fecha de diagnóstico
    # Ejemplo: si tienes una columna de fecha de diagnóstico en supledf
    if 'Fecha_Diagnostico' in df_.columns and 'FECHA_NAC' in df_.columns:
        
        df_['Edad Diagnóstico'] = df_.apply(
            lambda row: calcular_edad_diagnostico(row['FECHA_NAC'], row['Fecha_Diagnostico']), 
            axis=1
        )
        
        df_['Edad Diagnóstico (días)'] = df_.apply(
            lambda row: calcular_edad_diagnostico_dias(row['FECHA_NAC'], row['Fecha_Diagnostico']), 
            axis=1
        )
    #st.write(df_.shape)
    #st.dataframe(df_) 
    














    
    df_["RESUMEN"] = df_["Fecha_Diagnostico"].astype(str)+" "+df_["DIAGNOSTICO"]+" "+df_["ACTIVIDAD"]+" "+df_['Edad Diagnóstico']+" - "
    df_["ACTIVIDAD"] = df_["ACTIVIDAD"].str[4:]
    df_["ACTIVIDAD"] = df_["ACTIVIDAD"].str.strip()
    
    df_ = df_[df_["MICRORED"].notna()]
    
    df_["MICRORED"] = df_["MICRORED"]+" - "
    df_["ESTABLECIMIENTO"] = df_["ESTABLECIMIENTO"]+" - "
    df_["PACIENTE"] = df_["PACIENTE"]+" - "
    df_["Fecha_Diagnostico"] = df_["Fecha_Diagnostico"].astype(str)
    df_["Fecha_Diagnostico"] = df_["Fecha_Diagnostico"]+" - "
    df_["DIAGNOSTICO"] = df_["DIAGNOSTICO"]+" - "
    df_["Descripcion_Financiador"] = df_["Descripcion_Financiador"]+" - "
    df_["ACTIVIDAD"] = df_["ACTIVIDAD"]+" - "
    df_["Edad Diagnóstico (días)"] = df_["Edad Diagnóstico (días)"].astype(str)
    df_["Edad Diagnóstico"] = df_["Edad Diagnóstico"]+" - "
    
    dff = df_ = df_.groupby(["Establecimiento de Salud",'Documento', 'Periodos', '¿Es consecutivo?',"FECHA_NAC"])[
        ["MICRORED","ESTABLECIMIENTO","PACIENTE","Fecha_Diagnostico","DIAGNOSTICO","Descripcion_Financiador",
         "ACTIVIDAD","Edad Diagnóstico","RESUMEN"
         ]
    ].sum().reset_index()
    dff['Establecimiento de Salud'] = dff.apply(lambda x: tomar_ultimo_elemento(x['Establecimiento de Salud']),axis=1)
    dff['MICRORED'] = dff.apply(lambda x: eliminar_duplicados_col(x['MICRORED']),axis=1)
    dff['ESTABLECIMIENTO'] = dff.apply(lambda x: eliminar_duplicados_col(x['ESTABLECIMIENTO']),axis=1)
    dff['PACIENTE'] = dff.apply(lambda x: eliminar_duplicados_col(x['PACIENTE']),axis=1)
    #dff['Fecha_Diagnostico'] = dff.apply(lambda x: eliminar_duplicados_col(x['Fecha_Diagnostico']),axis=1)
    #dff['DIAGNOSTICO'] = dff.apply(lambda x: eliminar_duplicados_col(x['DIAGNOSTICO']),axis=1)
    dff['Descripcion_Financiador'] = dff.apply(lambda x: eliminar_duplicados_col(x['Descripcion_Financiador']),axis=1)
    #dff['ACTIVIDAD'] = dff.apply(lambda x: eliminar_duplicados_col(x['ACTIVIDAD']),axis=1)
    #dff['Edad Diagnóstico'] = dff.apply(lambda x: eliminar_duplicados_col(x['Edad Diagnóstico']),axis=1)
    
    dff["MICRORED"] = dff["MICRORED"].str[:-2]
    dff["ESTABLECIMIENTO"] = dff["ESTABLECIMIENTO"].str[:-2]
    #dff["Fecha_Diagnostico"] = dff["Fecha_Diagnostico"].str[:-2]
    #dff["DIAGNOSTICO"] = dff["DIAGNOSTICO"].str[:-2]
    dff["Descripcion_Financiador"] = dff["Descripcion_Financiador"].str[:-2]
    #dff["ACTIVIDAD"] = dff["ACTIVIDAD"].str[:-2]
    #dff["Edad Diagnóstico"] = dff["Edad Diagnóstico"].str[:-2]
    dff["PACIENTE"] = dff["PACIENTE"].str[:-2]
    print(dff.columns)

    # Contar la cantidad de diagnósticos por documento (solo una vez por fecha)
    # Primero obtenemos las fechas únicas por documento
    fechas_unicas_df = supledf.groupby(['Documento', 'Fecha_Diagnostico']).first().reset_index()
    # Luego contamos cuántas fechas únicas tiene cada documento
    cantidad_diagnosticos_df = fechas_unicas_df.groupby('Documento').agg({'Fecha_Diagnostico': 'count'}).reset_index()
    cantidad_diagnosticos_df = cantidad_diagnosticos_df.rename(columns={'Fecha_Diagnostico': 'Cantidad Diagnosticos'})
    
    # Agregar la columna del último diagnóstico y cantidad de diagnósticos al DataFrame final
    dff = pd.merge(dff, ultimo_diagnostico_df[['Documento', 'Ultimo Diagnostico']], on='Documento', how='left')
    dff = pd.merge(dff, cantidad_diagnosticos_df, on='Documento', how='left')
    dff['Ultimo Diagnostico'] = dff['Ultimo Diagnostico'].fillna('SIN DIAGNOSTICO')
    dff['Cantidad Diagnosticos'] = dff['Cantidad Diagnosticos'].fillna(0)

    dff = dff[['Establecimiento de Salud', 'Documento','PACIENTE', 'Periodos', '¿Es consecutivo?','FECHA_NAC', 'MICRORED', 'ESTABLECIMIENTO', 
            'Fecha_Diagnostico', 'DIAGNOSTICO', 'Descripcion_Financiador','ACTIVIDAD', 'Edad Diagnóstico','RESUMEN', 'Ultimo Diagnostico', 'Cantidad Diagnosticos']]
    dff = dff.sort_values(by="Establecimiento de Salud", ascending=True)
    #ANEMIA
    dff["CON ANEMIA?"] = dff["RESUMEN"].str.contains("ANEMIA", case=False, na=False).map({True: "TUVO ANEMIA", False: "NO TUVO ANEMIA"})

    # Separar la columna RESUMEN en varias columnas
    resumen_split = dff['RESUMEN'].str.split(' - ', expand=True)
    # Renombrar las columnas como RESUMEN_1, RESUMEN_2, ...
    resumen_split.columns = [f'ATENCIONES_{i+1}' for i in range(resumen_split.shape[1])]

    # Concatenar al DataFrame original (opcional: puedes eliminar la columna original si no la quieres)
    dff = pd.concat([dff.drop(columns=['RESUMEN']), resumen_split], axis=1)
    dff["Edad"] = dff["FECHA_NAC"].apply(calcular_edad)
    
    # Función para extraer meses totales del formato "X año(s), Y mes(es)"
    def extraer_meses_totales(edad_texto):
        try:
            if pd.isna(edad_texto) or edad_texto == "":
                return 0
            
            edad_str = str(edad_texto)
            meses_totales = 0
            
            # Extraer años
            if "año" in edad_str:
                anos_match = re.search(r'(\d+)\s*año', edad_str)
                if anos_match:
                    anos = int(anos_match.group(1))
                    meses_totales += anos * 12
            
            # Extraer meses
            if "mes" in edad_str:
                meses_match = re.search(r'(\d+)\s*mes', edad_str)
                if meses_match:
                    meses = int(meses_match.group(1))
                    meses_totales += meses
            
            return meses_totales
        except:
            return 0
    
    # Crear columna de clasificación por edad
    dff["Meses Totales"] = dff["Edad"].apply(extraer_meses_totales)
    dff["Clasificacion Edad"] = dff["Meses Totales"].apply(lambda x: "Mayor o igual a 12 meses" if x >= 12 else "Menor a 12 meses")
    num_cant = dff.shape[0]
    con_anemia_df = dff.groupby(["Ultimo Diagnostico"]).agg({"CON ANEMIA?": "count"}).reset_index()
    con_anemia_df = con_anemia_df.rename(columns={"CON ANEMIA?": "N° de Niños"})
    fig_con_anemia = px.bar(con_anemia_df,x="N° de Niños",y="Ultimo Diagnostico",title=f"Ultimo Diagnostico de Anemia ({num_cant} niños)",text="N° de Niños")
    fig_con_anemia.update_traces(textposition="outside", textfont_size=18, textfont_color="black")

    st.plotly_chart(fig_con_anemia)


    con_anemia_df_ = dff.groupby(["CON ANEMIA?"]).agg({"Edad": "count"}).reset_index()
    con_anemia_df_ = con_anemia_df_.sort_values(by="Edad",ascending=True)
    con_anemia_df_ = con_anemia_df_.rename(columns={"Edad": "N° de Niños"})
    fig_con_anemia_ = px.bar(con_anemia_df_,x="N° de Niños",y="CON ANEMIA?",title=f"",text="N° de Niños")#Anemia ({num_cant} niños)
    fig_con_anemia_.update_traces(textposition="outside", textfont_size=18, textfont_color="black")

    st.plotly_chart(fig_con_anemia_)

    # Gráfico de distribución de cantidad de diagnósticos (contando solo una vez por fecha) con clasificación por edad
    cantidad_diag_df = dff.groupby(["Cantidad Diagnosticos", "Clasificacion Edad"]).agg({"Edad": "count"}).reset_index()
    cantidad_diag_df = cantidad_diag_df.rename(columns={"Edad": "N° de Niños"})
    cantidad_diag_df["Cantidad Diagnosticos"] = cantidad_diag_df["Cantidad Diagnosticos"].astype(int).astype(str) + " diagnósticos"
    
    fig_cantidad_diag = px.bar(cantidad_diag_df, x="Cantidad Diagnosticos", y="N° de Niños", 
                              color="Clasificacion Edad",
                              title=f"Distribución de Fechas Únicas de Diagnóstico por Niño ({num_cant} niños)",
                              text="N° de Niños",
                              color_discrete_map={
                                  "Menor a 12 meses": "#FF6B6B",
                                  "Mayor o igual a 12 meses": "#4ECDC4"
                              })
    fig_cantidad_diag.update_traces(textposition="outside", textfont_size=14, textfont_color="black")
    fig_cantidad_diag.update_layout(
        xaxis_title="Cantidad de Diagnósticos", 
        yaxis_title="Número de Niños",
        legend=dict(
            orientation="v",
            yanchor="top",
            y=1,
            xanchor="left",
            x=1.02
        )
    )

    st.plotly_chart(fig_cantidad_diag)
    
    st.write(dff.shape)
    st.dataframe(dff)
    test_Dff = dff.groupby(["Cantidad Diagnosticos", "Clasificacion Edad","Ultimo Diagnostico"]).agg({"Edad": "count"}).reset_index()
    st.dataframe(test_Dff)
    st.download_button(
            label="Descargar",
            icon=":material/download:",
            data=convert_excel_df(dff),
            file_name=f"REDTRUJULLO_suplementacion_c1_2025.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
     )


    #st.dataframe(all_c1_carga_df)
    
    #padron_df = padron_df[
    #    [
    #        "Tipo_file","Tipo de Documento","Documento","DATOS NIÑO PADRON","DIRECCION PADRON","REFERENCIA DE DIRECCION","NUMERO DE DOCUMENTO  DE LA MADRE","DATOS MADRE PADRON","NUMERO DE CELULAR",
    #        "TIPO DE SEGURO","EESS NACIMIENTO","EESS","FRECUENCIA DE ATENCION"
    #    ]
    #]
    # dff = pd.merge(all_c1_carga_df,padron_df, left_on='Número de Documento del niño', right_on='Documento', how='left')
    #st.write(dff.shape)
    #st.dataframe(dff)
    #st.dataframe(dff)

def generar_excel_seguimiento_nominal():
    styles(2)
    c1_carga_df = fetch_carga_childs()
    
    
    padron_df = fetch_padron()
    padron_df = padron_df[
        [
            "Tipo_file","Tipo de Documento","Documento","DATOS NIÑO PADRON","DIRECCION PADRON","REFERENCIA DE DIRECCION","NUMERO DE DOCUMENTO  DE LA MADRE","DATOS MADRE PADRON","NUMERO DE CELULAR",
            "TIPO DE SEGURO","EESS NACIMIENTO","EESS","FRECUENCIA DE ATENCION"
        ]
    ]
    
    list_mes = [mes_short(x) for x in sorted(list(c1_carga_df["Mes"].unique()))]
    col1, col2 = st.columns(2)
    with col1:
        st.title("Manejador deSeguimiento Nominal")
    with col2:
        select_mes  = st.selectbox("Mes:",list_mes , key="select2",index=len(list_mes) - 1)
    mes = mestext_short(select_mes)
    seg_nominal_df =c1_carga_df[(c1_carga_df["Mes"] ==  mes) & (c1_carga_df["Año"] == 2025)]
    seg_nominal_df = seg_nominal_df[[
        "Establecimiento de Salud","Nombres del Actor Social","Tipo de Documento del niño","Número de Documento del niño","Fecha de Nacimiento","Rango de Edad","DNI de la madre","Celular de la madre",
        "Mes","Año"
    ]]
    dff = pd.merge(seg_nominal_df,padron_df, left_on='Número de Documento del niño', right_on='Documento', how='left')
    prioridad = {'DNI': 1, 'CUI': 2, 'CNV': 3}
    dff['Prioridad'] = dff['Tipo de Documento'].map(prioridad)
    dff = dff.sort_values(by=['Número de Documento del niño', 'Prioridad'])
    dff = dff.drop_duplicates(subset='Número de Documento del niño', keep='first')
    dff = dff.drop(columns=['Prioridad'])
    dff['Tipo_file'] = dff['Tipo_file'].fillna('Otro Ubigeo')
    pre_final_df = dff[[
        "Establecimiento de Salud","Nombres del Actor Social","Tipo de Documento","Número de Documento del niño","DATOS NIÑO PADRON",
        "Fecha de Nacimiento","Rango de Edad","DIRECCION PADRON","REFERENCIA DE DIRECCION","DNI de la madre","NUMERO DE DOCUMENTO  DE LA MADRE",
        "DATOS MADRE PADRON","NUMERO DE CELULAR","Celular de la madre","Tipo_file","TIPO DE SEGURO",'EESS NACIMIENTO', 'EESS', 'FRECUENCIA DE ATENCION'
    ]]
    pre_final_df["Número de Documento del niño"] = pre_final_df["Número de Documento del niño"].astype(str)
    nuevas_columnas = {
        'N°': pd.Series(dtype='int'),
        'Establecimiento de Atención': pd.Series(dtype='str'),
        '¿Es prematuro?': pd.Series(dtype='str'),
        'Fecha del tamizaje de Hemoglobina de 06 MESES': pd.Series(dtype='str'),
        'Resultado de Hemoglobina de 06 MESES': pd.Series(dtype='str'),
        '¿Está suplementado?': pd.Series(dtype='str'),
        'Tipo de SUPLEMENTO': pd.Series(dtype='str'),
        'Fecha del tamizaje de Hemoglobina de 07 MESES': pd.Series(dtype='str'),
        'Resultado de Hemoglobina de 07 MESES': pd.Series(dtype='str'),
        'Fecha del tamizaje de Hemoglobina de 08 MESES': pd.Series(dtype='str'),
        'Resultado de Hemoglobina de 08 MESES': pd.Series(dtype='str'),
        'Fecha del tamizaje de Hemoglobina de 09 MESES': pd.Series(dtype='str'),
        'Resultado de Hemoglobina de 09 MESES': pd.Series(dtype='str'),
        'Fecha del tamizaje de Hemoglobina de 12 MESES': pd.Series(dtype='str'),
        'Resultado de Hemoglobina de 12 MESES': pd.Series(dtype='str'),
        '¿Fue parte de una Sesion demostrativa?': pd.Series(dtype='str'),
        '¿Fue HISEADO (Sesión Demostrativa)?': pd.Series(dtype='str'),
        'Observaciones': pd.Series(dtype='str'),
    }
    #
    def obs_periodos_consecutivos(df):
        all_c1_carga_df = df[(df["Año"] == 2025)]
        all_c1_carga_df["Mes_name"] = all_c1_carga_df["Mes"].apply(mes_short)
        all_c1_carga_df["Mes_name"] = all_c1_carga_df["Mes_name"] + "-"
        all_c1_carga_df = all_c1_carga_df.sort_values(by="Mes", ascending=True)
        all_c1_carga_df["Mes"] = all_c1_carga_df["Mes"].astype(str)
        unique_childs25_df = all_c1_carga_df.groupby(["Número de Documento del niño"]).agg({"Mes": "sum","Mes_name": "sum"}).reset_index()
        unique_childs25_df.columns = ["Documento", "Periodos","Periodos_name"]
        unique_childs25_df["Periodos_name"] = unique_childs25_df["Periodos_name"].str[:-1]
        unique_childs25_df['Es_Consecutivo'] = unique_childs25_df['Periodos'].apply(es_consecutivo)
        unique_childs25_df = unique_childs25_df[["Documento","Periodos_name","Es_Consecutivo"]]
        unique_childs25_df.columns = ["Documento","Periodos","¿Es consecutivo?"]
        unique_childs25_df["Documento"] = unique_childs25_df["Documento"].astype(str)
        return unique_childs25_df
    periodos_consecutivos_df = obs_periodos_consecutivos(c1_carga_df)
    pre_final_df = pd.merge(pre_final_df, periodos_consecutivos_df, left_on="Número de Documento del niño", right_on="Documento", how="left")
    primer_dia_mes = datetime(int(2025), int(mes), 1)
    
    if int(mes) == 12:
            ultimo_dia_mes = datetime(int(2025) + 1, 1, 1) - pd.Timedelta(days=1)
    else:
            ultimo_dia_mes = datetime(int(2025), int(mes) + 1, 1) - pd.Timedelta(days=1)

    pre_final_df['Edad en días (primer día del mes)'] = pre_final_df['Fecha de Nacimiento'].apply(lambda x: (primer_dia_mes - x).days)
    pre_final_df['Edad en días (último día del mes)'] = pre_final_df['Fecha de Nacimiento'].apply(lambda x: (ultimo_dia_mes - x).days)

    pre_final_df['Niños 120-149 días en mes'] = pre_final_df.apply(
                lambda row: "SI" if (
                    (row['Edad en días (primer día del mes)'] <= 149 and row['Edad en días (último día del mes)'] >= 120)
                ) else "NO", axis=1
            )

    pre_final_df['Niños 180-209 días en mes'] = pre_final_df.apply(
                lambda row: "SI" if (
                    (row['Edad en días (primer día del mes)'] <= 209 and row['Edad en días (último día del mes)'] >= 180)
                ) else "NO", axis=1
            )

    pre_final_df['Niños 270-299 días en mes'] = pre_final_df.apply(
                lambda row: "SI" if (
                    (row['Edad en días (primer día del mes)'] <= 299 and row['Edad en días (último día del mes)'] >= 270)
                ) else "NO", axis=1
            )

    pre_final_df['Niños 360-389 días en mes'] = pre_final_df.apply(
                lambda row: "SI" if (
                    (row['Edad en días (primer día del mes)'] <= 389 and row['Edad en días (último día del mes)'] >= 360)
                ) else "NO", axis=1
            )
    pre_final_df['Rango de Días Activo'] = pre_final_df.apply(combinar_rangos_dias, axis=1)
    pre_final_df["Edad"] = pre_final_df['Fecha de Nacimiento'].apply(lambda x: calcular_edad(x))
    
    final_df = pre_final_df.assign(**nuevas_columnas)
    final_df = final_df[
        ['Establecimiento de Salud', 'Nombres del Actor Social','N°',
        'Tipo de Documento', 'Número de Documento del niño','DATOS NIÑO PADRON',
        'Fecha de Nacimiento', 'Rango de Edad','Edad',  'DNI de la madre','DATOS MADRE PADRON',
        'DIRECCION PADRON','REFERENCIA DE DIRECCION','NUMERO DE CELULAR',
        'TIPO DE SEGURO','Establecimiento de Atención','¿Es consecutivo?',
        'Rango de Días Activo','Periodos',
        '¿Es prematuro?',
        'Fecha del tamizaje de Hemoglobina de 06 MESES',
        'Resultado de Hemoglobina de 06 MESES',
        '¿Está suplementado?',
        'Tipo de SUPLEMENTO',
        'Fecha del tamizaje de Hemoglobina de 07 MESES',
        'Resultado de Hemoglobina de 07 MESES',
        'Fecha del tamizaje de Hemoglobina de 08 MESES',
        'Resultado de Hemoglobina de 08 MESES',
        'Fecha del tamizaje de Hemoglobina de 09 MESES',
        'Resultado de Hemoglobina de 09 MESES',
        'Fecha del tamizaje de Hemoglobina de 12 MESES',
        'Resultado de Hemoglobina de 12 MESES',
        '¿Fue parte de una Sesion demostrativa?',
        '¿Fue HISEADO (Sesión Demostrativa)?',
        'Observaciones'
    ]]
    final_df["Establecimiento de Salud"] = final_df["Establecimiento de Salud"].str.replace('LOS GRANADOS "SAGRADO CORAZON"', "LOS GRANADOS SAGRADO CORAZON")
    final_df = final_df.sort_values(by=["Establecimiento de Salud",'Nombres del Actor Social'])
    final_df["Número de Documento del niño"] = final_df["Número de Documento del niño"].astype(str)
    del pre_final_df
    # Crear el archivo Excel en memoria
    """
    AQUI EMPIEZA DATOS DE SEG JUNIO
    """
    segno_junio_df = pd.read_excel(r"C:\Proyectos\c1_vd\data\SEG_NOMINAL_JUNIO.xlsx")
    segno_junio_df['¿Es prematuro?'] = segno_junio_df['¿Es prematuro?'].fillna("NO")
    segno_junio_df['¿Es prematuro?'] = segno_junio_df['¿Es prematuro?'].str.strip()
    # Convertir todos los valores diferentes a "SI" en "NO"
    segno_junio_df['¿Es prematuro?'] = segno_junio_df['¿Es prematuro?'].apply(lambda x: "SI" if x == "SI" else "NO") 
    
    # Función para validar si un valor es numérico o string con números
    def es_valor_numerico(valor):
        if pd.isna(valor) or valor == "":
            return False
        if isinstance(valor, (int, float)):
            return True
        if isinstance(valor, str):
            # Verificar si el string contiene al menos un dígito
            return any(char.isdigit() for char in valor)
        return False
    
    # Aplicar la función a la columna de hemoglobina
    segno_junio_df['hemoglobina al mes prematuro'] = segno_junio_df['hemoglobina al mes prematuro'].apply(
        lambda x: x if es_valor_numerico(x) else pd.NA
    )
    
    # Función para validar si un valor es una fecha válida
    def es_fecha_valida(valor):
        if pd.isna(valor):
            return False
        if isinstance(valor, (pd.Timestamp, datetime)):
            return True
        if isinstance(valor, str):
            # Verificar si es un string que representa una fecha
            try:
                pd.to_datetime(valor, errors='raise')
                return True
            except:
                return False
        return False
    
    # Aplicar la función a la columna de fecha de tamizaje
    segno_junio_df['FECHA del tamizaje de Hemoglobina de 06 MESES'] = segno_junio_df['FECHA del tamizaje de Hemoglobina de 06 MESES'].apply(
        lambda x: x if es_fecha_valida(x) else pd.NA
    )
    
    # Función para validar si un valor es numérico (string, float o int)
    def es_numerico_puro(valor):
        if pd.isna(valor) or valor == "":
            return False
        if isinstance(valor, (int, float)):
            return True
        if isinstance(valor, str):
            # Verificar si el string contiene solo números, punto decimal y signos
            try:
                float(valor)
                return True
            except ValueError:
                return False
        return False
    
    # Aplicar la función a la columna de resultado de hemoglobina
    segno_junio_df['Resultado de Hemoglobina de 06 MESES'] = segno_junio_df['Resultado de Hemoglobina de 06 MESES'].apply(
        lambda x: x if es_numerico_puro(x) else pd.NA
    )
    segno_junio_df['¿Tiene ANEMIA? - de 10.5 a menos'] = segno_junio_df['¿Tiene ANEMIA? - de 10.5 a menos'].apply(lambda x: "SI" if x == "SI" else "NO") 
    segno_junio_df['¿Está suplementado?'] = segno_junio_df['¿Está suplementado?'].apply(lambda x: "SI" if x == "SI" else "NO") 
    segno_junio_df['Si estuvo en un cuadro de ANEMIA, y ya tiene 12 MESES: ¿Es un Niño recuperado que ya no tiene ANEMIA?'] = segno_junio_df['Si estuvo en un cuadro de ANEMIA, y ya tiene 12 MESES: ¿Es un Niño recuperado que ya no tiene ANEMIA?'].apply(lambda x: "SI" if x == "SI" else "NO") 
    segno_junio_df['¿Fue parte de una Sesion demostrativa?'] = segno_junio_df['¿Fue parte de una Sesion demostrativa?'].apply(lambda x: "SI" if x == "SI" else "NO")
    
    print(segno_junio_df.columns)
    segno_junio_df = segno_junio_df.drop(columns=['Unnamed: 0'])
    columns_to_keep = ['Documento del Niño', 'Apellico y Nombre del Niño',
       'Periodos Cargados',  '¿Es prematuro?',
       'hemoglobina al mes prematuro',
       'FECHA del tamizaje de Hemoglobina de 06 MESES',
       'Resultado de Hemoglobina de 06 MESES',
       '¿Tiene ANEMIA? - de 10.5 a menos', '¿Está suplementado?',
       'Tipo de SUPLEMENTO', '07 MESES: Fecha y resultado de Hemoglobina',
       '8 MESES: Fecha y resultado de Hemoglobina',
       '09 MESES: Fecha y resultado de Hemoglobina',
       '10 MESES: Fecha y resultado de Hemoglobina',
       '11 MESES: Fecha y resultado de Hemoglobina',
       '12 MESES: Fecha y resultado de Hemoglobina',
       'Si estuvo en un cuadro de ANEMIA, y ya tiene 12 MESES: ¿Es un Niño recuperado que ya no tiene ANEMIA?',
       '¿Fue parte de una Sesion demostrativa?', 'TIPO SEGURO',
       'Observaciones', 'CONSUME HIERRO  DE 4 A 5 MESES', 'EESS ATENCION',
       'HG', 'SUPLEMENTO DE 4 A 5 MESES', 'EESS ATENCIONS',
       '¿Fue HISEADO(Tamizaje 6 meses)?',
       '08 MESES: Fecha y resultado de Hemoglobina', '¿Fue HISEADO(SD)?',
       '09 MESES:  Fecha y resultado de Hemoglobina',
       'ESTABLECIMIENTO DE SALUD ']
    segno_junio_df = segno_junio_df[columns_to_keep]
     
     # Crear un diccionario para renombrar las columnas agregando _1
    rename_dict = {}
    for col in segno_junio_df.columns:
         if col not in ['Documento del Niño']:  # No renombrar la columna de join
             rename_dict[col] = col + '_1'
     
     # Renombrar las columnas
    segno_junio_df = segno_junio_df.rename(columns=rename_dict)
    segno_junio_df = segno_junio_df.rename(columns={"Documento del Niño": "Número de Documento del niño"})
     # Hacer el join entre final_df y segno_junio_df
    segno_junio_df["Número de Documento del niño"] = segno_junio_df["Número de Documento del niño"].astype(str)
    
    final_df = pd.merge(final_df, segno_junio_df, 
                       on='Número de Documento del niño', 
                        how='left')
    #final_df['¿Es prematuro?'] = final_df['¿Es prematuro?'].fillna(final_df['¿Es prematuro?_1'])
    #final_df['Fecha del tamizaje de Hemoglobina de 06 MESES '] = final_df['Fecha del tamizaje de Hemoglobina de 06 MESES '].fillna(final_df['Fecha del tamizaje de Hemoglobina de 06 MESES'])
    
     # Completar espacios vacíos usando los datos de segno_junio_df
     # Para cada columna que existe en ambos dataframes
    
    columns_to_fill = [
         ('¿Es prematuro?', '¿Es prematuro?_1'),
         ('Fecha del tamizaje de Hemoglobina de 06 MESES', 'FECHA del tamizaje de Hemoglobina de 06 MESES_1'),
         ('Resultado de Hemoglobina de 06 MESES', 'Resultado de Hemoglobina de 06 MESES_1'),
         ('¿Está suplementado?', '¿Está suplementado?_1'),
         ('Tipo de SUPLEMENTO', 'Tipo de SUPLEMENTO_1'),
         ('¿Fue parte de una Sesion demostrativa?', '¿Fue parte de una Sesion demostrativa?_1'),
         ('Fecha del tamizaje de Hemoglobina de 07 MESES', '07 MESES: Fecha y resultado de Hemoglobina_1'),
         ('Fecha del tamizaje de Hemoglobina de 08 MESES', '08 MESES: Fecha y resultado de Hemoglobina_1'),
         ('Fecha del tamizaje de Hemoglobina de 09 MESES', '09 MESES: Fecha y resultado de Hemoglobina_1'),
         ('Fecha del tamizaje de Hemoglobina de 12 MESES', '12 MESES: Fecha y resultado de Hemoglobina_1'),
         ('Observaciones', 'Observaciones_1'),
     ]
     
     # Llenar valores nulos en final_df con valores de segno_junio_df
    for col_final, col_junio in columns_to_fill:
         if col_final in final_df.columns and col_junio in final_df.columns:
             final_df[col_final] = final_df[col_final].fillna(final_df[col_junio])
     
     # Opcional: eliminar las columnas _1 después de completar los datos
    cols_to_drop = [col for col in final_df.columns if col.endswith('_1') or col == 'Documento del Niño']
    
    final_df = final_df.drop(columns=cols_to_drop, errors='ignore')
    final_df.to_excel(r"C:\Proyectos\c1_vd\data\seguimiento_nominal_por_establecimiento_julio.xlsx")
    st.dataframe(final_df)
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Crear hoja de resumen SIN formato ni tabla
        resumen = pd.DataFrame({
            'Establecimiento de Salud': final_df['Establecimiento de Salud'].unique(),
            'Total de Registros': [len(final_df[final_df['Establecimiento de Salud'] == est]) 
                                for est in final_df['Establecimiento de Salud'].unique()]
        })
        resumen = resumen.sort_values(by='Total de Registros',ascending=False)
        resumen.to_excel(writer, sheet_name='Resumen', index=False)
        
        # Crear hojas por establecimiento CON formato de tabla
        establecimientos = final_df['Establecimiento de Salud'].unique()
        for establecimiento in establecimientos:
            df_establecimiento = final_df[final_df['Establecimiento de Salud'] == establecimiento]
            df_establecimiento['N°'] = range(1, len(df_establecimiento) + 1)
            nombre_hoja = str(establecimiento)[:31]
            df_establecimiento.to_excel(writer, sheet_name=nombre_hoja, index=False)
            
            worksheet = writer.sheets[nombre_hoja]
            max_row = len(df_establecimiento) + 1  # +1 para encabezado
            max_col = len(df_establecimiento.columns)
            table_range = f'A1:{get_column_letter(max_col)}{max_row}'
            table_name = f"Table_{nombre_hoja.replace(' ', '_').replace('-', '_')}"
            
            # Solo agregar la tabla si hay datos
            if len(df_establecimiento) > 0:
                table = Table(displayName=table_name, ref=table_range)
                style = TableStyleInfo(
                    name="TableStyleMedium9",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False
                )
                table.tableStyleInfo = style
                worksheet.add_table(table)
            
            # Ajustar ancho de columnas
            for idx, col in enumerate(df_establecimiento.columns):
                max_length = max(
                    df_establecimiento[col].astype(str).apply(len).max(),
                    len(col)
                )
                worksheet.column_dimensions[get_column_letter(idx + 1)].width = max_length + 2
            
            worksheet.freeze_panes = 'A2'
            # Desbloquear todas las celdas primero
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.protection = Protection(locked=False)
            
            # Bloquear la fila de encabezados (nombres de columnas) y aplicar formato
            for cell in worksheet[1]:
                cell.protection = Protection(locked=True)
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
            worksheet.row_dimensions[1].height = 40
            
            # Proteger solo la columna "Número de Documento del niño"
            col_to_lock_name = 'Número de Documento del niño'
            header_row = worksheet[1]
            col_to_lock_idx = None
            for idx, cell in enumerate(header_row, 1):
                if cell.value == col_to_lock_name:
                    col_to_lock_idx = idx
                    break
            if col_to_lock_idx:
                for row in range(2, max_row + 1):  # Empezar desde la fila 2 (después del encabezado)
                    cell = worksheet.cell(row=row, column=col_to_lock_idx)
                    cell.protection = Protection(locked=True)
            # Habilitar la protección de la hoja permitiendo la mayoría de las acciones
            worksheet.protection.sheet = True
            worksheet.protection.sort = False
            worksheet.protection.autoFilter = False
            worksheet.protection.formatCells = False
            worksheet.protection.formatColumns = False
            worksheet.protection.formatRows = False
    
    # Obtener los datos del archivo Excel
    excel_data = output.getvalue()
    
    st.download_button(
            label="Descargar Reporte Seguimiento Nominal",
            icon=":material/download:",
            data=excel_data,
            file_name=f"seguimiento_nominal_por_establecimiento_{select_mes}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
     )
    st.write(final_df.shape)
    
    st.dataframe(final_df)
    


"""

"""




def hb_data_c1():
    styles(2)
    st.title("Datos de Hemoglobina")
    if "dataframe_childs"  in st.session_state:
        df = st.session_state["dataframe_childs"]
        print(df.columns)
        df = df.rename(columns={"Número de Documento": "DNI_PACIENTE"})
        df['Rango de Días Activo'] = df.apply(combinar_rangos_dias, axis=1)
        df = df.drop(columns=[
            'Edad Dias','Edad en días (primer día del mes)','Edad en días (último día del mes)', 'Niños 120-149 días en mes','Niños 180-209 días en mes', 'Niños 270-299 días en mes',
        'Niños 360-389 días en mes','Estado Padrón Nominal','Tipo Registro Padrón Nominal','Entidad Actualiza', 'FECHA DE MODIFICACIÓN DEL REGISTRO',
        'USUARIO QUE MODIFICA', 'FRECUENCIA DE ATENCION', 'EESS ADSCRIPCIÓN','EESS ULTIMA ATENCION', 'Fecha Ultima Atención', 'Zona', 'Manzana','Sector','Número Doc Jefe Familia(P)', 'Datos Jefe Famlia(P)', 
        'Total de Intervenciones', 'Total de VD presenciales Válidas','Total de VD presencial Válidas WEB','Total de VD presencial Válidas MOVIL', 'Tipo Documento',
        'MENOR VISITADO','¿MENOR ENCONTRADO?','Estado Visitas'
       ])
        
        hb_df = pd.read_excel(r"./data/microred/DOSAJES DE HEMOGLOBINA NIÑOS DE 6 MESES A 1 AÑO_17_07_2025.xlsx",sheet_name="BASE")
        hb_df["Fecha_Atencion"] = pd.to_datetime(hb_df["Fecha_Atencion"]).dt.date
        hb_df["DNI_PACIENTE"] = hb_df["DNI_PACIENTE"].astype(str).str.strip()
        hb_df["Hemoglobina"] = hb_df["Hemoglobina"].fillna(0)
        hb_df["Resultados"] = hb_df["Fecha_Atencion"].astype(str) + " - " + hb_df["Hemoglobina"].astype(str) + " | "
        hbdff = hb_df.groupby(["DNI_PACIENTE"]).agg({"Resultados": "sum"}).reset_index()
        
        dataframe_final = pd.merge(df,hbdff,on="DNI_PACIENTE",how="left")

        # Aplicar la función para crear las nuevas columnas
        dataframe_final[['Ultima Fecha tamizaje', 'Ultima HB']] = dataframe_final['Resultados'].apply(
            lambda x: pd.Series(extraer_ultimo_resultado(x))
        )
        # Crear la columna ESTADO TAMIZAJE
        dataframe_final['Edad Ultimo Tamizaje'] = dataframe_final.apply(
            lambda row: calcular_edad_diagnostico(row['Fecha de Nacimiento'], row['Ultima Fecha tamizaje']), 
            axis=1
        )
        dataframe_final['ESTADO TAMIZAJE'] = dataframe_final['Ultima HB'].apply(determinar_estado_tamizaje)
        
        # Opcional: Eliminar la columna Resultados original si no la necesitas
        # dataframe_final = dataframe_final.drop(columns=['Resultados'])
        
        st.write(dataframe_final.shape)
        st.dataframe(dataframe_final)
        st.download_button(
            label="Descargar Reporte HB",
            icon=":material/download:",
            data=convert_excel_df(dataframe_final),
            file_name=f"Seguimiento_hb_julio_2025_MPT.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    else:
        st.warning("Ingrese a Vistas a Niños")


def c1_2025_avances():
    styles(2)
    st.title("Avances C1 2025")
    df1 = pd.read_excel(r"C:\Proyectos\c1_vd\data\2025\vd_niños_junio_2025_corte_final.xlsx")
    df2 = pd.read_excel(r"C:\Proyectos\c1_vd\data\2025\vd_niños_julio_2025_corte_final.xlsx")
    
    dff = pd.concat([df1,df2])
    #dff = dff.sort_values(by="Mes",ascending=True)
    dff["Mes_"] = dff["Mes"].map({6: "Junio", 7: "Julio"})
    dff["Rango de Edad"] = dff["Rango de Edad"].str.replace("3-5 meses","1-5 meses")
    print(dff.columns)
    st.dataframe(dff)
    cg_mes_df =dff.groupby(["Mes","Mes_","Rango de Edad"])[["Tipo Documento"]].count().reset_index()
    fig_1 = px.bar(cg_mes_df,x="Mes_",y="Tipo Documento",color="Rango de Edad",title="",barmode="group",text_auto=True)
    fig_1.update_traces(textfont_size=20, textangle=0, textposition="outside", cliponaxis=False)
    fig_1.update_layout(legend_font_size=20,legend_title_text="",legend=dict(
        orientation="h",
        yanchor="bottom",
        y=1.02,
        xanchor="right",
        x=1
    ))
    fig_1.update_layout(
        title=dict(
            text=""
        ),
        xaxis=dict(
            title=dict(
                text="MES"
            )
        ),
        yaxis=dict(
            title=dict(
                text="N° de Niños Cargados"
            )
        ),
        
        #font=dict(
        #    family="Courier New, monospace",
        #    size=18,
        #    color="RebeccaPurple"
        #)
    )
    st.plotly_chart(fig_1)
    
    
    
    
    

def wwww():
    styles(2)
    st.title("TWWWWW")
    childs_df = pd.read_parquet(r"./data/backups/carga_nino.parquet")
    childs_df = childs_df.sort_values(by=["Mes"])
    #unique_childs25_df = all_c1_carga_df.groupby(["Número de Documento del niño"]).agg({"Mes": "sum","Mes_name": "sum","Establecimiento de Salud": "sum"}).reset_index()
    #unique_childs25_df.columns = ["Documento", "Periodos","Periodos_name","Establecimiento de Salud"]
    #unique_childs25_df["Periodos_name"] = unique_childs25_df["Periodos_name"].str[:-1] 
    #unique_childs25_df['Es_Consecutivo'] = unique_childs25_df['Periodos'].apply(es_consecutivo)  
    childs_df["Mes"] = childs_df["Mes"].astype(str) + "-"
    childs_df["Establecimiento de Salud"] = childs_df["Establecimiento de Salud"]+ " - "
    childs_df = childs_df.groupby(["Número de Documento del niño"]).agg({"Mes": "sum","Establecimiento de Salud": "sum"}).reset_index()
    
    childs_df["Mes"] = childs_df["Mes"].str[:-1]
    childs_df["Periodos"] = childs_df["Mes"].str.replace("-","")
    childs_df["Periodos"] = childs_df["Periodos"].astype(int)
    childs_df['Establecimiento de Salud'] = childs_df.apply(lambda x: tomar_ultimo_elemento(x['Establecimiento de Salud']),axis=1)
    def es_consecutivo(numero):
        s = str(numero)
        
        def get_sequences(text):
            if not text:
                return [[]]
            sequences = []
            # Months can be 1 or 2 digits (1-12)
            # Check for 1-digit month
            if len(text) >= 1:
                num = int(text[:1])
                if 1 <= num <= 12:
                    for seq in get_sequences(text[1:]):
                        sequences.append([num] + seq)
            # Check for 2-digit month
            if len(text) >= 2:
                num = int(text[:2])
                if 10 <= num <= 12:
                    for seq in get_sequences(text[2:]):
                        sequences.append([num] + seq)
            return sequences

        possible_sequences = get_sequences(s)
        
        for seq in possible_sequences:
            if len(seq) > 1:
                # Check for any consecutive pair in the sequence
                for i in range(1, len(seq)):
                    if seq[i] - seq[i-1] == 1:
                        return "Consecutivo"
                    
        return "No Consecutivo"

    childs_df['Es_Consecutivo'] = childs_df['Periodos'].apply(es_consecutivo)
    childs_df = childs_df.rename(columns={"Número de Documento del niño":"DNI_PACIENTE"})
    childs_df["DNI_PACIENTE"] = childs_df["DNI_PACIENTE"].astype(str)
    st.write(childs_df.shape)
    st.dataframe(childs_df)
    diag_df = pd.read_parquet("Diagnosticos.parquet")
    #print(diag_df.columns)
    #print(diag_df.info())
    #st.write(diag_df.shape)
    today = pd.to_datetime('today')
    diag_df['Edad_Actual_Dias'] = (today - pd.to_datetime(diag_df['FECHA_NAC'])).dt.days
    diag_df = diag_df.sort_values(by=["Edad_Actual_Dias"],ascending=False)
    def calculate_age_years_months(birth_date, diag_date):
        birth_date = pd.to_datetime(birth_date)
        diag_date = pd.to_datetime(diag_date)

        total_months = (diag_date.year - birth_date.year) * 12 + (diag_date.month - birth_date.month)
        
        if diag_date.day < birth_date.day:
            total_months -= 1

        years = total_months // 12
        months = total_months % 12
        
        return f"{years} años,{months} meses"

    diag_df['Edad_Diagnostico'] = diag_df.apply(lambda row: calculate_age_years_months(row['FECHA_NAC'], row['Fecha_Diagnostico']), axis=1)
    
    diag_df["Diag_corto"] = diag_df["DIAGNOSTICO"]
    diag_df["Diag_corto"] = diag_df["Diag_corto"].replace(
        {"SUPLEMENTACIÓN CON HIERRO": "SH",
        "SUPLEMENTACIÓN DE MULTIMICRONUTRIENTES":"SMU",
        "Dosaje de Hemoglobina con hemoglobinómetro":"DHMO",
        "Dosaje de Hemoglobina":"DH",
        "ANEMIA POR DEFICIENCIA DE HIERRO SIN ESPECIFICACION":"ASH",
        "ANEMIA DE TIPO NO ESPECIFICADO":"ANOES"
        }
    )
    #st.dataframe(diag_df)
    doc_unicos_diag = diag_df["DNI_PACIENTE"].unique()
    ## agrupando
    diag_df["Diag_corto"] = diag_df["Diag_corto"]+" - "
    diag_df["Edad_Diagnostico"] = diag_df["Edad_Diagnostico"]+" - "
    diag_df["Anemia"] = (diag_df["Diag_corto"].astype(str) + diag_df["Edad_Diagnostico"].astype(str)).where(
        diag_df["Diag_corto"].str.contains("A", na=False),
        ""
    )
    diag_df["Anemia"] = diag_df["Anemia"].str[:-3]
    diag_df["Anemia"] = diag_df["Anemia"].str.replace("ASH","")
    diag_df["Anemia"] = diag_df["Anemia"].str.replace("ANOES","")
    
    st.dataframe(diag_df)
    diag_df["FECHA_NAC"] = diag_df["FECHA_NAC"].astype(str)
    diag_df_rs = diag_df.groupby(["DNI_PACIENTE","FECHA_NAC","Edad_Actual_Dias"])[["Diag_corto","Edad_Diagnostico","Anemia"]].sum().reset_index()
    diag_df_rs["DNI_PACIENTE"] = diag_df_rs["DNI_PACIENTE"].astype(str)
    st.write(f"DOC_UNICOS:{len(doc_unicos_diag)}")
    st.write(diag_df_rs.shape)
    st.dataframe(diag_df_rs)

    ### JOIN
    childs_df = childs_df.merge(diag_df_rs, on="DNI_PACIENTE", how="left")
    # Rellenos por defecto para evaluar toda la población
    childs_df["Diag_corto"] = childs_df["Diag_corto"].fillna("-")
    childs_df["Edad_Diagnostico"] = childs_df["Edad_Diagnostico"].fillna("-")
    childs_df["Anemia"] = childs_df["Anemia"].fillna("-")
    ####FILKTROS (removido para evaluar toda la población)
    #childs_df = childs_df[childs_df["Es_Consecutivo"]=="Consecutivo"]
    #####################
    childs_df['Edad_Diagnostico'] = childs_df.apply(lambda x: eliminar_duplicados_col(x['Edad_Diagnostico']),axis=1)
    # Quitar solo un " - " al final si existe, sin truncar texto (p.ej. "meses")
    childs_df['Edad_Diagnostico'] = childs_df['Edad_Diagnostico'].str.replace(r"\s-\s$", '', regex=True)
    # Valor por defecto "-" si queda vacío
    childs_df['Edad_Diagnostico'] = childs_df['Edad_Diagnostico'].where(childs_df['Edad_Diagnostico'].str.strip() != '', '-')
    #childs_df["Edad_Diagnostico"] = childs_df["Edad_Diagnostico"].fillna("XXXXX")
    childs_df['Anemia'] = childs_df.apply(lambda x: eliminar_duplicados_col(x['Anemia']),axis=1)
    # Quitar solo un " - " al inicio si existe, sin truncar texto
    childs_df["Anemia"] = childs_df["Anemia"].str.replace(r"^\s-\s", '', regex=True)
    # Valor por defecto "-" si queda vacío
    childs_df['Anemia'] = childs_df['Anemia'].where(childs_df['Anemia'].str.strip() != '', '-')
    # Crear columnas únicas por cada edad de diagnóstico, contando presencia por fila
    # Divide por " - " y genera columnas con 1/0 según aparición
    # Unificar separador con regex para tolerar espacios variables
    edades_list = childs_df['Edad_Diagnostico'].fillna('').str.split(r'\s*-\s*')
    edades_exploded = edades_list.explode()
    edades_exploded = edades_exploded.dropna()
    edades_exploded = edades_exploded[edades_exploded.str.strip() != '']
    edades_dummies = pd.crosstab(edades_exploded.index, edades_exploded)
    childs_df = childs_df.join(edades_dummies)
    # Si no hubo valores (None/""), rellena con 0 en columnas creadas
    childs_df[edades_dummies.columns] = childs_df[edades_dummies.columns].fillna(0).astype(int)
    # Crear columnas únicas por cada edad en Anemia, con prefijo "A - "
    anemia_list = childs_df['Anemia'].fillna('').str.split(r'\s*-\s*')
    anemia_exploded = anemia_list.explode()
    anemia_exploded = anemia_exploded.dropna()
    anemia_exploded = anemia_exploded[anemia_exploded.str.strip() != '']
    anemia_dummies = pd.crosstab(anemia_exploded.index, anemia_exploded)
    anemia_dummies = anemia_dummies.rename(columns=lambda c: f"A - {c}")
    childs_df = childs_df.join(anemia_dummies)
    # Rellenar con 0 en filas sin valores de Anemia
    childs_df[anemia_dummies.columns] = childs_df[anemia_dummies.columns].fillna(0).astype(int)
    childs_df["Atenciones antes 12 meses"] = childs_df[['0 años,6 meses', '0 años,7 meses','0 años,8 meses', '0 años,9 meses','0 años,10 meses','0 años,11 meses']].sum(axis=1)
    def condi_val1(x):
        if x >= 1:
            return 1
        else:
            return 0
    childs_df["1 años,0 meses_"] = childs_df["1 años,0 meses"].apply(condi_val1)
    childs_df["Condicion 2 Atenciones"] = childs_df["Atenciones antes 12 meses"]+childs_df["1 años,0 meses_"]
    
    # Condición: >= 2 -> "Cumple", si no -> "No Cumple". NaN se asume 0
    def condi2_(x,mes12):
        if x >= 2 and mes12 == 1:
            return "Cumple"
        else:
            return "No Cumple"
    childs_df["Condicion 2 Atenciones_"] = childs_df.apply(lambda x: condi2_(x["Condicion 2 Atenciones"],x["1 años,0 meses_"]),axis=1)
    childs_df = childs_df.drop(columns=["1 años,0 meses_"])
    # Calcular edad máxima al 31/12/2025 en días basada en FECHA_NAC
    fecha_corte_2025 = pd.Timestamp('2025-12-31')
    nac_dt = pd.to_datetime(childs_df["FECHA_NAC"], errors="coerce")
    childs_df["MAXIMA EDAD 2025"] = (fecha_corte_2025 - nac_dt).dt.days
    # Manejo seguro: si falta fecha o es posterior al corte, colocar 0
    childs_df["MAXIMA EDAD 2025"] = childs_df["MAXIMA EDAD 2025"].clip(lower=0).fillna(0).astype(int)
    print(childs_df.columns)
    st.write(childs_df.shape)
    st.dataframe(childs_df)
    childs_df.to_parquet("avances_tamizajes.parquet",index=False,engine = "pyarrow")
    st.success("Archivo guardado con éxito")
    #print(diag_df["DIAGNOSTICO"].unique())
    #print(diag_df["ACTIVIDAD"].unique())




def buscar_sector_childs():
    styles(2)
    st.title("Sectorizar")
    childs_df = pd.read_parquet(r"./data/backups/carga_nino.parquet")
    childs_df = childs_df[childs_df["Año"]==2025]
    childs_df["Dirección"] = childs_df["Dirección"].str.strip()
    childs_df["Dirección"] = childs_df["Dirección"].str.replace(r'[^\w\s]', '', regex=True)
    print(childs_df.columns)
    direccion_input = st.text_input("Dirección")
    if direccion_input:
        childs_df = childs_df[childs_df["Dirección"].str.contains(direccion_input, case=False, na=False)]
    childs_df = childs_df.groupby(['Nombres del Actor Social','Dirección','Zona', 'Manzana', 'Sector','Establecimiento de Salud']).size().reset_index(name='count')
    childs_df = childs_df[[
        'Nombres del Actor Social','Dirección','Zona', 'Manzana', 'Sector','Establecimiento de Salud'
    ]]
    
    st.subheader(f"Registros: {childs_df.shape[0]}")
    st.dataframe(childs_df)
    